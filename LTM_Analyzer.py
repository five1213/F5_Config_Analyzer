import configparser
import datetime
import json
import os
import re
import time

import pandas as pd
from openpyxl import load_workbook  # 导入openpyxl

result_path = 'E:\\result\\'

config_path = 'C:\\Users\\Administrator\\Desktop\\nas'

analyzer_path = 'C:\\Users\\Administrator\\Desktop\\NAS配置文件设备列表.xlsx'

config = configparser.ConfigParser()
config.read('config/config.ini', encoding='utf-8')

port_json_str_open = open('config/port.json', encoding='utf-8' ,errors='ignore')
port_json_str = port_json_str_open.read()

ports_dir = json.loads(port_json_str)
ports_data = dict([val, key] for key, val in ports_dir.items())

port_json_str_open.close()

nas_fliename_list_path = 'C:\\Users\\Administrator\\Desktop\\'
NAS_FILENAME = config.get('NAS', 'NAS_FILENAME')
LTM_V12_ROUTE_RE_STR = config.get('LTM', 'LTM_V12_ROUTE_RE_STR')
LTM_V10_ROUTE_RE_STR = config.get('LTM', 'LTM_V10_ROUTE_RE_STR')
LTM_V12_SELF_IP_RE_STR = config.get('LTM', 'LTM_V12_SELF_IP_RE_STR')
LTM_V10_SELF_IP_RE_STR = config.get('LTM', 'LTM_V10_SELF_IP_RE_STR')
LTM_HTTP_ACL_RE_STR = config.get('LTM', 'LTM_HTTP_ACL_RE_STR')
LTM_SSH_ACL_RE_STR = config.get('LTM', 'LTM_SSH_ACL_RE_STR')
NSAE_IP_RE_STR = config.get('LTM', 'NSAE_IP_RE_STR')
NSAE_ROUTE_RE_STR = config.get('LTM', 'NSAE_ROUTE_RE_STR')
NSAE_HTTP_ACL_RE_STR = config.get('LTM', 'NSAE_HTTP_ACL_RE_STR')
NSAE_SSH_ACL_RE_STR = config.get('LTM', 'NSAE_SSH_ACL_RE_STR')

NSAE_SLB_REAL_RE_STR = config.get('LTM', 'NSAE_SLB_REAL_RE_STR')
NSAE_SLB_REAL_DISABLE_RE_STR = config.get('LTM', 'NSAE_SLB_REAL_DISABLE_RE_STR')
NSAE_SLB_GROUP_MEMBER_RE_STR = config.get('LTM', 'NSAE_SLB_GROUP_MEMBER_RE_STR')
NSAE_SLB_VIRTUAL_RE_STR = config.get('LTM', 'NSAE_SLB_VIRTUAL_RE_STR')
NSAE_SLB_POLICY_RE_STR = config.get('LTM', 'NSAE_SLB_POLICY_RE_STR')
NSAE_SSL_HOST_RE_STR = config.get('LTM', 'NSAE_SSL_HOST_RE_STR')

CITRIX_IP_RE_STR = config.get('LTM', 'CITRIX_IP_RE_STR')
CITRIX_ROUTE_RE_STR = config.get('LTM', 'CITRIX_ROUTE_RE_STR')
CITRIX_ACL_RE_STR = config.get('LTM', 'CITRIX_ACL_RE_STR')

LTM_V12_SSL_CERT_RE_STR = config.get('LTM', 'LTM_V12_SSL_CERT_RE_STR')
LTM_V12_SSL_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_SSL_PROFILE_RE_STR')
LTM_V12_SSL_VS_RE_STR = config.get('LTM', 'LTM_V12_SSL_VS_RE_STR')

LTM_V12_SOURCE_PERSIST_RE_STR = config.get('LTM', 'LTM_V12_SOURCE_PERSIST_RE_STR')
LTM_V12_COOKIE_PERSIST_RE_STR = config.get('LTM', 'LTM_V12_COOKIE_PERSIST_RE_STR')
LTM_V12_HTTP_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_HTTP_PROFILE_RE_STR')
LTM_V12_TCP_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_TCP_PROFILE_RE_STR')
LTM_V12_FASTL4_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_FASTL4_PROFILE_RE_STR')
LTM_V12_POOL_RE_STR = config.get('LTM', 'LTM_V12_POOL_RE_STR')
LTM_V12_POOL_MEMBER_RE_STR = config.get('LTM', 'LTM_V12_POOL_MEMBER_RE_STR')
LTM_V12_VS_RE_STR = config.get('LTM', 'LTM_V12_VS_RE_STR')

LTM_V12_SSL_CERT_EXP_RE_STR = config.get('LTM', 'LTM_V12_SSL_CERT_EXP_RE_STR')
LTM_V12_SSL_PROFILE_EXP_RE_STR = config.get('LTM', 'LTM_V12_SSL_PROFILE_EXP_RE_STR')

nas_filename_pattern = re.compile(NAS_FILENAME, re.MULTILINE)
ltm_v12_route_pattern = re.compile(LTM_V12_ROUTE_RE_STR, re.MULTILINE)
ltm_v10_route_pattern = re.compile(LTM_V10_ROUTE_RE_STR, re.MULTILINE)
ltm_v12_self_ip_pattern = re.compile(LTM_V12_SELF_IP_RE_STR, re.MULTILINE)
ltm_v10_self_ip_pattern = re.compile(LTM_V10_SELF_IP_RE_STR, re.MULTILINE)
ltm_http_acl_pattern = re.compile(LTM_HTTP_ACL_RE_STR, re.MULTILINE)
ltm_ssh_acl_pattern = re.compile(LTM_SSH_ACL_RE_STR, re.MULTILINE)
nsae_ip_pattern = re.compile(NSAE_IP_RE_STR, re.MULTILINE)
nsae_route_pattern = re.compile(NSAE_ROUTE_RE_STR, re.MULTILINE)
nsae_http_acl_pattern = re.compile(NSAE_HTTP_ACL_RE_STR, re.MULTILINE)
nsae_ssh_acl_pattern = re.compile(NSAE_SSH_ACL_RE_STR, re.MULTILINE)

nsae_slb_real_pattern = re.compile(NSAE_SLB_REAL_RE_STR, re.MULTILINE)
nsae_slb_real_disable_pattern = re.compile(NSAE_SLB_REAL_DISABLE_RE_STR, re.MULTILINE)
nsae_slb_group_member_pattern = re.compile(NSAE_SLB_GROUP_MEMBER_RE_STR, re.MULTILINE)
nsae_slb_virtual_pattern = re.compile(NSAE_SLB_VIRTUAL_RE_STR, re.MULTILINE)
nsae_slb_policy_pattern = re.compile(NSAE_SLB_POLICY_RE_STR, re.MULTILINE)
nsae_ssl_host_pattern = re.compile(NSAE_SSL_HOST_RE_STR, re.MULTILINE)

citrix_ip_pattern = re.compile(CITRIX_IP_RE_STR, re.MULTILINE)
citrix_route_pattern = re.compile(CITRIX_ROUTE_RE_STR, re.MULTILINE)
citrix_acl_pattern = re.compile(CITRIX_ACL_RE_STR, re.MULTILINE)

ltm_v12_ssl_cert_pattern = re.compile(LTM_V12_SSL_CERT_RE_STR, re.MULTILINE)
ltm_v12_ssl_profile_pattern = re.compile(LTM_V12_SSL_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_ssl_vs_pattern = re.compile(LTM_V12_SSL_VS_RE_STR, re.MULTILINE)

ltm_v12_source_persist_pattern = re.compile(LTM_V12_SOURCE_PERSIST_RE_STR, re.MULTILINE)
ltm_v12_cookie_persist_pattern = re.compile(LTM_V12_COOKIE_PERSIST_RE_STR, re.MULTILINE)
ltm_v12_http_profile_pattern = re.compile(LTM_V12_HTTP_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_tcp_profile_pattern = re.compile(LTM_V12_TCP_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_fastl4_profile_pattern = re.compile(LTM_V12_FASTL4_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_pool_pattern = re.compile(LTM_V12_POOL_RE_STR, re.MULTILINE)
ltm_v12_pool_member_pattern = re.compile(LTM_V12_POOL_MEMBER_RE_STR, re.MULTILINE)
ltm_v12_vs_pattern = re.compile(LTM_V12_VS_RE_STR, re.MULTILINE)

ltm_v12_ssl_cert_exp_pattern = re.compile(LTM_V12_SSL_CERT_EXP_RE_STR, re.MULTILINE)
ltm_v12_ssl_profile_exp_pattern = re.compile(LTM_V12_SSL_PROFILE_EXP_RE_STR, re.MULTILINE)


ls_config_path = os.listdir(config_path)

NAS_DEVICE_DIR = {}

def get_nas_device_list():
    nas_fliename_list = []
    for name in ls_config_path:
        filename = nas_filename_pattern.findall(name)
        if filename[0].find('FW') < 0 and filename[0].find('fw') < 0 and filename[0].find('TAP') < 0 and filename[0].find('tap') < 0 and filename[0].find('RT') < 0:
            nas_fliename_list.append(filename[0])

    df = pd.DataFrame(nas_fliename_list, columns=['设备名称'])
    df.to_excel(nas_fliename_list_path+'NAS配置文件设备列表.xlsx', index=False)

def get_nas_filename_list():
    nas_fliename_list = {}
    for name in ls_config_path:
        filename = nas_filename_pattern.findall(name)
        nas_fliename_list[filename[0]] = name
    return nas_fliename_list

def get_device_data(path):
    wb = load_workbook(path)  # 打开Excel
    sheet = wb['NAS配置文件列表']  # 定位表单
    for row in range(2, sheet.max_row + 1):
        device = {}
        name = sheet.cell(row, 1).value.strip()
        device['name'] = name
        device['type']  = sheet.cell(row, 2).value
        device['version'] = sheet.cell(row, 3).value
        device['real_name'] = sheet.cell(row, 4).value
        device['mgmt_ip'] = sheet.cell(row, 5).value
        NAS_DEVICE_DIR[name] = device

    ltm_device_list = []
    sheet2 = wb['LTM解析设备列表']
    for row in range(2, sheet2.max_row + 1):
        ltm_device_list.append(sheet2.cell(row, 1).value)
    return ltm_device_list



def get_ltm_config(filepath,type,version):

    ltm_config_open = open(filepath, encoding='utf-8' ,errors='ignore')
    ltm_config_open_str = ltm_config_open.read()
    ltm_config_open.close()

    ltm_v12_source_persist_map = {}
    ltm_v12_source_persist_map['source_addr'] = '3600'
    ltm_v12_source_persist = ltm_v12_source_persist_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_source_persist:
        name = item[0].strip()
        time_out = item[1].strip()
        ltm_v12_source_persist_map[name] = time_out

    ltm_v12_cookie_persist_map = {}
    ltm_v12_cookie_persist_map['cookie'] = '::encrypt:disabled::name:null::method:insert::'
    ltm_v12_cookie_persist = ltm_v12_cookie_persist_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_cookie_persist:
        name = item[0].strip()
        is_encrypt = item[1].strip()
        cookie_name = item[2].strip()
        method = item[3].strip()
        ltm_v12_cookie_persist_map[name] = '::encrypt:'+is_encrypt+'::name:'+cookie_name+'::method:'+method+'::'

    ltm_v12_http_profile_map = {}
    ltm_v12_http_profile_map['http'] = 'disabled'
    ltm_v12_http_profile = ltm_v12_http_profile_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_http_profile:
        name = item[0].strip()
        xforwarded = item[1].strip()
        ltm_v12_http_profile_map[name] = xforwarded

    ltm_v12_tcp_profile_map = {}
    ltm_v12_tcp_profile_map['tcp'] = '300'
    ltm_v12_tcp_profile = ltm_v12_tcp_profile_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_tcp_profile:
        name = item[0].strip()
        idle_timeout = item[1].strip()
        ltm_v12_tcp_profile_map[name] = idle_timeout

    ltm_v12_fastl4_profile_map = {}
    ltm_v12_fastl4_profile_map['fastL4'] = '::timeout:300::pva:full::'
    ltm_v12_fastl4_profile = ltm_v12_fastl4_profile_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_fastl4_profile:
        name = item[0].strip()
        idle_timeout = item[1].strip()
        pva = item[2].strip()
        ltm_v12_fastl4_profile_map[name] = '::timeout:'+idle_timeout+'::pva:'+pva+'::'

    ltm_v12_pool_map = {}
    ltm_v12_pool = ltm_v12_pool_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_pool:
        name = item[0].strip()
        balanc_mode = item[1].strip()
        members_str = item[2].strip()
        members_info = 'none'
        if members_str != 'none':
            ltm_v12_pool_member = ltm_v12_pool_member_pattern.findall(members_str)
            members_info_detail = ''
            members_info_simple = ''
            for item2 in ltm_v12_pool_member:
                ip_port_str = item2[0].strip()
                ip_port_info = ''
                if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$",ip_port_str):
                    ipports = ip_port_str.split(":")
                    ip = ipports[0]
                    port = ipports[1]
                    if port in ports_data.keys():
                        port = ports_data[port]
                    ip_port_info = ip + ":" + port
                else:
                    ipports = ip_port_str.split(".")
                    ip = ipports[0]
                    port = ipports[1]
                    if port in ports_data.keys():
                        port = ports_data[port]
                    ip_port_info = ip + "." + port
                con_limit = item2[1].strip()
                dynamic_ratio = item2[2].strip()
                mem_monitor = item2[3].strip()
                priority = item2[4].strip()
                ratio = item2[5].strip()
                session = item2[6].strip()
                state = item2[7].strip()
                members_info_detail = members_info_detail + ip_port_info + ' ' + session + ' ' + state + ' l:' + con_limit + ' p:' + priority + ' r:' + ratio  + '\n'
                if session == 'user-enabled' or session == 'monitor-enabled':
                    members_info_simple  = members_info_simple + ip_port_info + '\n'

                members_info = '##members_info_simple#' + members_info_simple + '##members_info_detail#' + members_info_detail + '##'

        monitor = item[3].strip()
        ltm_v12_pool_map[name] = '::balanc_mode:'+balanc_mode+'::monitor:'+monitor+'::members_info:'+members_info+'::'

    ltm_v12_vs_list = []
    ltm_v12_vs = ltm_v12_vs_pattern.findall(ltm_config_open_str)

    for item in ltm_v12_vs:
        vs = ['']*28
        name = item[0].strip()
        vs[0] = name
        vs_conn_limit = item[1].strip()
        vs[1] = vs_conn_limit
        vs_ip_port_str = item[2].strip()
        vs_ip_port_info = ''
        if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$", vs_ip_port_str):
            ipports = vs_ip_port_str.split(":")
            ip = ipports[0]
            port = ipports[1]
            if port in ports_data.keys():
                port = ports_data[port]
            vs_ip_port_info = ip + ":" + port
        else:
            ipports = vs_ip_port_str.split(".")
            ip = ipports[0]
            port = ipports[1]
            if port in ports_data.keys():
                port = ports_data[port]
            vs_ip_port_info = ip + "." + port
        vs[2] = vs_ip_port_info
        vs_status = item[3].strip()
        vs[3] = vs_status
        vs_protocol = item[4].strip()
        vs[4] = vs_protocol
        vs_persist_str = item[5].strip()
        vs_persist_name = 'none'
        vs_persist_mothod = 'none'
        vs_persist_timeout = ''
        persist_cookie_encrypt = ''
        persist_cookie_name = ''
        persist_cookie_method = ''
        if vs_persist_str != 'none':
            vs_persist_str_pattern = re.compile("{\s*([\s\S]*?)\s{\n", re.MULTILINE)
            vs_persist_name = ''.join(vs_persist_str_pattern.findall(vs_persist_str))
            if vs_persist_name in ltm_v12_source_persist_map.keys():
                vs_persist_mothod = 'source_addr'
                vs_persist_timeout = ltm_v12_source_persist_map[vs_persist_name]
            elif vs_persist_name in ltm_v12_cookie_persist_map.keys():
                vs_persist_mothod = 'session_cookie'
                cookie_persist_str = ltm_v12_cookie_persist_map[vs_persist_name]
                cookie_encrypt_pattern = re.compile("::encrypt:([\s\S]*?)::", re.MULTILINE)
                persist_cookie_encrypt = ''.join(cookie_encrypt_pattern.findall(cookie_persist_str))
                cookie_name_pattern = re.compile("::name:([\s\S]*?)::", re.MULTILINE)
                persist_cookie_name = ''.join(cookie_name_pattern.findall(cookie_persist_str))
                cookie_method_pattern = re.compile("::method:([\s\S]*?)::", re.MULTILINE)
                persist_cookie_method = ''.join(cookie_method_pattern.findall(cookie_persist_str))

        vs[5] = vs_persist_name
        vs[6] = vs_persist_mothod
        vs[7] = vs_persist_timeout
        vs[8] = persist_cookie_encrypt
        vs[9] = persist_cookie_name
        vs[10] = persist_cookie_method

        vs_pool = item[6].strip()
        vs_pool_name = 'none'
        vs_balanc_mode = ''
        vs_pool_monitor = ''
        members_info_simple = ''
        members_info_detail = ''
        if vs_pool != 'none':
            vs_pool_name = vs_pool
            if vs_pool_name in ltm_v12_pool_map.keys():
                vs_pool_info_str = ltm_v12_pool_map[vs_pool_name]
                vs_balanc_mode_pattern = re.compile("::balanc_mode:([\s\S]*?)::", re.MULTILINE)
                vs_balanc_mode = ''.join(vs_balanc_mode_pattern.findall(vs_pool_info_str))
                vs_pool_monitor_pattern = re.compile("::monitor:([\s\S]*?)::", re.MULTILINE)
                vs_pool_monitor = ''.join(vs_pool_monitor_pattern.findall(vs_pool_info_str))
                vs_members_info_pattern = re.compile("::members_info:([\s\S]*?)::", re.MULTILINE)
                vs_members_info = ''.join(vs_members_info_pattern.findall(vs_pool_info_str))
                if vs_members_info != 'none':
                    members_info_simple_pattern = re.compile("##members_info_simple#([\s\S]*?)##", re.MULTILINE)
                    members_info_simple = ''.join(members_info_simple_pattern.findall(vs_pool_info_str))
                    members_info_detail_pattern = re.compile("##members_info_detail#([\s\S]*?)##", re.MULTILINE)
                    members_info_detail = ''.join(members_info_detail_pattern.findall(vs_pool_info_str))

        vs[11] = vs_pool_name
        vs[12] = vs_balanc_mode
        vs[13] = vs_pool_monitor
        vs[14] = members_info_simple.strip('\n')
        vs[15] = members_info_detail.strip('\n')

        vs_profiles = item[7].strip().strip('{').strip('}')
        fastl4_profile_name = ''
        fastl4_timeout = ''
        fastl4_pva = ''
        tcp_profile_name = ''
        tcp_profile_timeout = ''
        http_profile_name = ''
        http_profile_xforwarded = ''
        other_profile = ''
        profiles_info_pattern = re.compile("^\s*([\s\S]*?)\s{\n\s*context[\s\S]*?}", re.MULTILINE)
        profiles_list = profiles_info_pattern.findall(vs_profiles)
        for profile in profiles_list:
            profile_name = profile.strip()
            if profile_name in ltm_v12_fastl4_profile_map.keys():
                fastl4_profile_name = profile_name
                fastl4_info_str = ltm_v12_fastl4_profile_map[profile_name]
                fastl4_timeout_pattern = re.compile("::timeout:([\s\S]*?)::", re.MULTILINE)
                fastl4_timeout = ''.join(fastl4_timeout_pattern.findall(fastl4_info_str))
                fastl4_pva_pattern = re.compile("::pva:([\s\S]*?)::", re.MULTILINE)
                fastl4_pva = ''.join(fastl4_pva_pattern.findall(fastl4_info_str))
            elif profile_name in ltm_v12_tcp_profile_map.keys():
                tcp_profile_name = profile_name
                tcp_profile_timeout = ltm_v12_tcp_profile_map[profile_name]
            elif profile_name in ltm_v12_http_profile_map.keys():
                http_profile_name = profile_name
                http_profile_xforwarded = ltm_v12_http_profile_map[profile_name]
            else:
                other_profile = other_profile + '\n' + profile_name

        vs[16] = fastl4_profile_name
        vs[17] = fastl4_timeout
        vs[18] = fastl4_pva
        vs[19] = tcp_profile_name
        vs[20] = tcp_profile_timeout
        vs[21] = http_profile_name
        vs[22] = http_profile_xforwarded
        vs[23] = other_profile.strip('\n')
        vs_rules = item[8].strip()
        vs[24] = vs_rules
        vs_snat_pool_str = item[9].strip()
        vs_snat_pool_name = 'none'
        if vs_snat_pool_str != 'none':
            vs_snat_pool_pattern = re.compile("\s*pool\s([\s\S]*?)\n", re.MULTILINE)
            vs_snat_pool_name = ''.join(vs_snat_pool_pattern.findall(vs_snat_pool_str))

        vs[25] = vs_snat_pool_name
        vs_source_port = item[10].strip()
        vs[26] = vs_source_port
        vs_vlans = item[11].strip()
        vs[27] = vs_vlans
        ltm_v12_vs_list.append(vs)

    return ltm_v12_vs_list


def get_ssl_exp_config(filepath,type,version):

    ssl_config_open = open(filepath, encoding='utf-8' ,errors='ignore')
    ssl_config_open_str = ssl_config_open.read()
    ltm_v12_ssl_cert_map = {}

    ltm_v12_ssl_cert_map['default.crt'] = "::expir:" + '2099-12-31 22:00:00' + '::domain:' + 'f5.com'+ '::'
    end_time = '2022-01-03 12:00:00'

    timeArray = time.strptime(end_time, "%Y-%m-%d %H:%M:%S")

    end_time_stamp =int(time.mktime(timeArray))

    ltm_v12_ssl_exp_cert_map = {}
    ltm_v12_ssl_cert = ltm_v12_ssl_cert_exp_pattern.findall(ssl_config_open_str)
    for item in ltm_v12_ssl_cert:
        name = item[0].strip()
        expir_date = int(item[1].strip())
        expir_date_int = expir_date

        expir_date_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(expir_date_int))

        cert_info = item[2].strip()
        cert_cn_pattern = re.compile("[\s\S]*?CN=([\s\S]*?),[\s\S]*?", re.MULTILINE)
        cert_cn = ''.join(cert_cn_pattern.findall(cert_info))
        dns_info = item[3].strip()
        if dns_info != 'none':
            dns_info_pattern = re.compile("DNS:([\s\S]*?)[,|\"]", re.MULTILINE)
            dns_names = dns_info_pattern.findall(dns_info)
            if len(dns_names) != 0:
                dnss = ''
                for dns_name in dns_names:
                    dnss = dnss + dns_name + '\n'
                cert_cn = dnss.rstrip('\n')
            else:
                cert_cn = dns_info.split(':')[1].strip()

        ltm_v12_ssl_cert_map[name] = "::expir:" + expir_date_str + '::domain:' + cert_cn.strip()+ '::'
        if expir_date <= end_time_stamp:
            ltm_v12_ssl_exp_cert_map[name] = "::expir:" + expir_date_str + '::domain:' + cert_cn.strip()+ '::'

    ltm_v12_ssl_profile_map = {}
    ltm_v12_ssl_profile_exp_map = {}
    ltm_v12_ssl_profile = ltm_v12_ssl_profile_exp_pattern.findall(ssl_config_open_str)
    for item in ltm_v12_ssl_profile:
        name = item[0].strip()
        profile_cert = item[1].strip()
        profile_chain_cert = item[2].strip()
        profile_ca_cert = item[4].strip()
        profile_cert_key = item[5].strip()

        if profile_cert in ltm_v12_ssl_exp_cert_map.keys():
            if name in ltm_v12_ssl_profile_exp_map:
                ltm_v12_ssl_profile_exp_map[name] = ltm_v12_ssl_profile_exp_map[name] + "::cert:" + profile_cert + "::key:" + profile_cert_key + "::"
            else:
                ltm_v12_ssl_profile_exp_map[name] = "::cert:" + profile_cert + "::key:" + profile_cert_key + "::"

        if profile_chain_cert in ltm_v12_ssl_exp_cert_map.keys():
            if name in ltm_v12_ssl_profile_exp_map:
                ltm_v12_ssl_profile_exp_map[name] = ltm_v12_ssl_profile_exp_map[name] + "::chain:" + profile_chain_cert + "::"
            else:
                ltm_v12_ssl_profile_exp_map[name] = "::chain:" +profile_chain_cert + "::"

        if profile_ca_cert in ltm_v12_ssl_exp_cert_map.keys():
            if name in ltm_v12_ssl_profile_exp_map:
                ltm_v12_ssl_profile_exp_map[name] = ltm_v12_ssl_profile_exp_map[name] + "::ca:" + profile_ca_cert + "::"
            else:
                ltm_v12_ssl_profile_exp_map[name] = "::ca:" + profile_ca_cert + "::"

        cert_cn = ltm_v12_ssl_cert_map[profile_cert]
        ltm_v12_ssl_profile_map[name] = cert_cn

    ltm_v12_ssl_vs_list = []
    ltm_v12_ssl_exp_in_vs_list = []
    ltm_v12_ssl_vs = ltm_v12_ssl_vs_pattern.findall(ssl_config_open_str)
    for item in ltm_v12_ssl_vs:
        ltm_v12_ssl_vs_info = [''] * 4
        vs_name = item[0].strip()
        ltm_v12_ssl_vs_info[0] = vs_name
        vs_ip_port = item[1].strip()
        ltm_v12_ssl_vs_info[1] = vs_ip_port
        profiles_info = item[2]
        profiles_info_pattern = re.compile("^\s*([\s\S]*?)\s{\n\s*context[\s\S]*?}", re.MULTILINE)
        profiles_list = profiles_info_pattern.findall(profiles_info)
        for profile in profiles_list:
            profile_name = profile.strip()
            if profile_name in ltm_v12_ssl_profile_map.keys():
                ltm_v12_ssl_vs_info[2] = profile_name
                ltm_v12_ssl_vs_info[3] = ltm_v12_ssl_profile_map[profile_name]
                if profile_name in ltm_v12_ssl_profile_exp_map.keys():
                    exp_in_vs = [''] * 8
                    exp_in_vs[0] = vs_name
                    exp_in_vs[1] = profile_name
                    exp_cert_info = ltm_v12_ssl_profile_exp_map.pop(profile_name)
                    cert_pattern = re.compile("::cert:([\s\S]*?)::", re.MULTILINE)
                    cert_name = ''.join(cert_pattern.findall(exp_cert_info))
                    exp_in_vs[2] = cert_name
                    cert_key_pattern = re.compile("::key:([\s\S]*?)::", re.MULTILINE)
                    cert_key_name = ''.join(cert_key_pattern.findall(exp_cert_info))
                    exp_in_vs[3] = cert_key_name
                    cert_chain_pattern = re.compile("::chain:([\s\S]*?)::", re.MULTILINE)
                    cert_chain_name = ''.join(cert_chain_pattern.findall(exp_cert_info))
                    exp_in_vs[4] = cert_chain_name
                    cert_ca_pattern = re.compile("::ca:([\s\S]*?)::", re.MULTILINE)
                    cert_ca_name = ''.join(cert_ca_pattern.findall(exp_cert_info))
                    exp_in_vs[5] = cert_ca_name
                    cert_expir_domain =  ltm_v12_ssl_cert_map[cert_name]
                    cert_domain_pattern = re.compile("::domain:([\s\S]*?)::", re.MULTILINE)
                    cert_domain = ''.join(cert_domain_pattern.findall(cert_expir_domain))
                    exp_in_vs[6] = cert_domain
                    cert_expir_pattern = re.compile("::expir:([\s\S]*?)::", re.MULTILINE)
                    cert_expir = ''.join(cert_expir_pattern.findall(cert_expir_domain))
                    exp_in_vs[7] = cert_expir
                    ltm_v12_ssl_exp_in_vs_list.append(exp_in_vs)

        ltm_v12_ssl_vs_list.append(ltm_v12_ssl_vs_info)

    ltm_v12_ssl_exp_cert_profile_list = []
    for profile in ltm_v12_ssl_profile_exp_map.keys():
        ltm_v12_ssl_exp_cert_info = [''] * 7
        ltm_v12_ssl_exp_cert_info[0] =  '' + profile
        exp_cert_info = ltm_v12_ssl_profile_exp_map[profile]
        cert_pattern = re.compile("::cert:([\s\S]*?)::", re.MULTILINE)
        cert_name = ''.join(cert_pattern.findall(exp_cert_info))
        ltm_v12_ssl_exp_cert_info[1] = cert_name
        cert_key_pattern = re.compile("::key:([\s\S]*?)::", re.MULTILINE)
        cert_key_name = ''.join(cert_key_pattern.findall(exp_cert_info))
        ltm_v12_ssl_exp_cert_info[2] = cert_key_name
        cert_chain_pattern = re.compile("::chain:([\s\S]*?)::", re.MULTILINE)
        cert_chain_name = ''.join(cert_chain_pattern.findall(exp_cert_info))
        ltm_v12_ssl_exp_cert_info[3]= cert_chain_name
        cert_ca_pattern = re.compile("::ca:([\s\S]*?)::", re.MULTILINE)
        cert_ca_name = ''.join(cert_ca_pattern.findall(exp_cert_info))
        ltm_v12_ssl_exp_cert_info[4]= cert_ca_name
        cert_expir_domain = ltm_v12_ssl_cert_map[cert_name]
        cert_domain_pattern = re.compile("::domain:([\s\S]*?)::", re.MULTILINE)
        cert_domain = ''.join(cert_domain_pattern.findall(cert_expir_domain))
        ltm_v12_ssl_exp_cert_info[5] = cert_domain
        cert_expir_pattern = re.compile("::expir:([\s\S]*?)::", re.MULTILINE)
        cert_expir = ''.join(cert_expir_pattern.findall(cert_expir_domain))
        ltm_v12_ssl_exp_cert_info[6] = cert_expir

        ltm_v12_ssl_exp_cert_profile_list.append(ltm_v12_ssl_exp_cert_info)

    ltm_v12_ssl_exp_cert_list = []
    for cert_exp_name in ltm_v12_ssl_exp_cert_map.keys():
        cert_exp_name_info = [''] * 3
        cert_exp_name_info[0] = cert_exp_name
        cert_expir_domain = ltm_v12_ssl_exp_cert_map[cert_exp_name]
        cert_domain_pattern = re.compile("::domain:([\s\S]*?)::", re.MULTILINE)
        cert_domain = ''.join(cert_domain_pattern.findall(cert_expir_domain))
        cert_exp_name_info[1] = cert_domain
        cert_expir_pattern = re.compile("::expir:([\s\S]*?)::", re.MULTILINE)
        cert_expir = ''.join(cert_expir_pattern.findall(cert_expir_domain))
        cert_exp_name_info[2] = cert_expir

        ltm_v12_ssl_exp_cert_list.append(cert_exp_name_info)

    ssl_exp_info_map = {}
    ssl_exp_info_map['ssl_exp_can_del_info'] = ltm_v12_ssl_exp_cert_profile_list
    ssl_exp_info_map['ssl_exp_not_del_info'] = ltm_v12_ssl_exp_in_vs_list
    ssl_exp_info_map['ssl_exp_cert_info'] = ltm_v12_ssl_exp_cert_list


    ssl_config_open.close()

    return ssl_exp_info_map

def get_ssl_config(filepath,type,version):

    ssl_config_open = open(filepath, encoding='utf-8' ,errors='ignore')
    ssl_config_open_str = ssl_config_open.read()
    ltm_v12_ssl_cert_map = {}
    ltm_v12_ssl_cert_map['default.crt'] = "::expir:" + '2099-12-31 22:00:00' + '::domain:' + 'f5.com'+ '::'

    ltm_v12_ssl_cert = ltm_v12_ssl_cert_pattern.findall(ssl_config_open_str)
    for item in ltm_v12_ssl_cert:
        name = item[0].strip()
        expir_date = int(item[1].strip())
        expir_date_str = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(expir_date))
        cert_info = item[2].strip()
        cert_cn_pattern = re.compile("[\s\S]*?CN=([\s\S]*?),[\s\S]*?", re.MULTILINE)
        cert_cn = ''.join(cert_cn_pattern.findall(cert_info))
        dns_info = item[3].strip()
        if dns_info != 'none':
            dns_info_pattern = re.compile("DNS:([\s\S]*?)[,|\"]", re.MULTILINE)
            dns_names = dns_info_pattern.findall(dns_info)
            if len(dns_names) != 0:
                dnss = ''
                for dns_name in dns_names:
                    dnss = dnss + dns_name + '\n'
                cert_cn = dnss.rstrip('\n')
            else:
                cert_cn = dns_info.split(':')[1].strip()

        ltm_v12_ssl_cert_map[name] =  "::expir:" + expir_date_str + '::domain:' + cert_cn.strip() + '::'

    ltm_v12_ssl_profile_map = {}
    ltm_v12_ssl_profile = ltm_v12_ssl_profile_pattern.findall(ssl_config_open_str)
    for item in ltm_v12_ssl_profile:
        name = item[0].strip()
        profile_cert = item[1].strip()
        cert_cn = ltm_v12_ssl_cert_map[profile_cert]
        ltm_v12_ssl_profile_map[name] = cert_cn

    ltm_v12_ssl_vs_list = []
    ltm_v12_ssl_vs = ltm_v12_ssl_vs_pattern.findall(ssl_config_open_str)
    for item in ltm_v12_ssl_vs:
        ltm_v12_ssl_vs_info = [''] * 5
        vs_name = item[0].strip()
        ltm_v12_ssl_vs_info[0] = vs_name
        vs_ip_port_str = item[1].strip()
        vs_ip_port_info = ''
        if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$", vs_ip_port_str):
            ipports = vs_ip_port_str.split(":")
            ip = ipports[0]
            port = ipports[1]
            if port in ports_data.keys():
                port = ports_data[port]
            vs_ip_port_info = ip + ":" + port
        else:
            ipports = vs_ip_port_str.split(".")
            ip = ipports[0]
            port = ipports[1]
            if port in ports_data.keys():
                port = ports_data[port]
            vs_ip_port_info = ip + "." + port
        ltm_v12_ssl_vs_info[1] = vs_ip_port_info
        profiles_info = item[2]
        profiles_info_pattern = re.compile("^\s*([\s\S]*?)\s{\n\s*context[\s\S]*?}", re.MULTILINE)
        profiles_list = profiles_info_pattern.findall(profiles_info)
        for profile in profiles_list:
            profile_name = profile.strip()
            if profile_name in ltm_v12_ssl_profile_map.keys():
                ltm_v12_ssl_vs_info[2] = profile_name
                cert_expir_domain = ltm_v12_ssl_profile_map[profile_name]
                cert_domain_pattern = re.compile("::domain:([\s\S]*?)::", re.MULTILINE)
                cert_domain = ''.join(cert_domain_pattern.findall(cert_expir_domain))
                ltm_v12_ssl_vs_info[3] = cert_domain
                cert_expir_pattern = re.compile("::expir:([\s\S]*?)::", re.MULTILINE)
                cert_expir = ''.join(cert_expir_pattern.findall(cert_expir_domain))
                ltm_v12_ssl_vs_info[4] = cert_expir
                break
        ltm_v12_ssl_vs_list.append(ltm_v12_ssl_vs_info)

    ssl_config_open.close()
    return ltm_v12_ssl_vs_list

def get_ltm_base_config(filepath,type,version):
    ltm_config = {}
    ltm_config_open = open(filepath, encoding='utf-8' ,errors='ignore')
    ltm_config_open_str = ltm_config_open.read()
    routes = ''
    self_ips = ''
    float_ips = ''
    acls = ''

    if type == 'F5':
        if version == 'V12' or version == 'V11' or version == 'V13':
            ltm_v12_route = ltm_v12_route_pattern.findall(ltm_config_open_str)
            for item in ltm_v12_route:
                network = item[1].strip()
                gw = item[0].strip()
                routes = routes + network + ' gw ' + gw + '\n'

            ltm_v12_self_ip = ltm_v12_self_ip_pattern.findall(ltm_config_open_str)
            for item in ltm_v12_self_ip:
                ip = item[0].strip()
                is_float = item[1].strip()
                traffic_group = item[2].strip()
                vlan = item[3].strip()
                if is_float == 'enabled':
                    float_ips = float_ips + ip + ' ' + traffic_group + '\n'
                else:
                    self_ips = self_ips + ip + ' ' + vlan + '\n'

        elif version == 'V10':
            ltm_v10_route = ltm_v10_route_pattern.findall(ltm_config_open_str)
            for item in ltm_v10_route:
                network = item[0].strip()
                gw = item[1].strip()
                routes = routes + network + ' gw ' + gw + '\n'

            ltm_v10_self_ip = ltm_v10_self_ip_pattern.findall(ltm_config_open_str)
            for item in ltm_v10_self_ip:
                ip = item[0].strip()
                is_float = item[1].strip()
                unit = item[2].strip()
                vlan = item[3].strip()
                if is_float == 'enabled':
                    float_ips = float_ips + ip + '\n'
                else:
                    self_ips = self_ips + ip + ' ' + vlan + '\n'

        ltm_http_acl = ltm_http_acl_pattern.findall(ltm_config_open_str)
        https_acls = ltm_http_acl[0].strip().replace(' ','\n')

        ltm_ssh_acl = ltm_ssh_acl_pattern.findall(ltm_config_open_str)
        ssh_acls = ltm_ssh_acl[0].strip().replace(' ','\n')

        acls = 'https_acl:\n' + https_acls + '\nssh_acl:\n' + ssh_acls

    elif type == 'NSAE':
        nsae_ip = nsae_ip_pattern.findall(ltm_config_open_str)
        for item in nsae_ip:
            interface = item[0].strip().replace('"','')
            ip = item[1].strip()
            mask = item[2].strip()
            self_ips = self_ips + ip + '/' + mask + ' ' + interface + '\n'

        nsae_route = nsae_route_pattern.findall(ltm_config_open_str)
        for item in nsae_route:
            type = item[0].strip()
            route = item[1].strip()
            if type == 'default':
                routes = routes + 'default' + ' gw ' + route + '\n'
            elif type == 'static':
                patterns = r' +'
                network = re.split(patterns,route)
                routes = routes + network[0] + '/' + network[1] + ' gw ' + network[2] + '\n'

        nsae_http_acl = nsae_http_acl_pattern.findall(ltm_config_open_str)
        http_acls = ''
        for item in nsae_http_acl:
            acl_temp = item.strip().replace(' ','/')
            http_acls = http_acls + acl_temp + '\n'

        nsae_ssh_acl = nsae_ssh_acl_pattern.findall(ltm_config_open_str)
        ssh_acls = ''
        for item in nsae_ssh_acl:
            acl_temp = item.strip().replace(' ','/')
            ssh_acls = ssh_acls + acl_temp + '\n'

        acls = 'https_acl:\n' + http_acls + 'ssh_acl:\n' + ssh_acls


    elif type == 'Citrix':
        citrix_ip = citrix_ip_pattern.findall(ltm_config_open_str)
        for item in citrix_ip:
            vlan_id = item[0].strip()
            ip = item[1].strip()
            mask = item[2].strip()
            self_ips = self_ips + ip + '/' + mask + ' vlan ' + vlan_id + '\n'
            float_ips = float_ips + ip + '/' + mask + ' vlan ' + vlan_id + '\n'

        citrix_route = citrix_route_pattern.findall(ltm_config_open_str)
        for item in citrix_route:
            network = item[0].strip()
            mask = item[1].strip()
            gw = item[2].strip()
            routes = routes + network + '/' + mask + ' gw ' + gw + '\n'

        citrix_acl = citrix_acl_pattern.findall(ltm_config_open_str)
        for item in citrix_acl:
            acl_allow_src = item[0].strip()
            acl_allow_dest = item[1].strip()
            acls = acls + 'src: ' + acl_allow_src + ' dest: ' + acl_allow_dest + '\n'

    ltm_config['route'] = routes.rstrip('\n')
    ltm_config['self_ip'] = self_ips.rstrip('\n')
    ltm_config['float_ip'] = float_ips.rstrip('\n')
    ltm_config['acls'] = acls.rstrip('\n')
    ltm_config_open.close()
    return ltm_config

def get_nsae_ssl_config(filepath,type,version):
    ssl_config_open = open(filepath, encoding='utf-8' ,errors='ignore')
    ssl_config_open_str = ssl_config_open.read()
    nsae_slb_real_map = {}
    nsae_slb_real_list = nsae_slb_real_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_real_list:
        nsae_slb_real_name = re.sub('\s*','',item[0])
        nsae_slb_real_ip = re.sub('\s*','',item[1])
        nsae_slb_real_port = re.sub('\s*','',item[2])
        nsae_slb_real_limit = re.sub('\s*','',item[3])
        nsae_slb_real_check = re.sub('\s*','',item[4])
        nsae_slb_real_map[nsae_slb_real_name] = "##ipport#" + nsae_slb_real_ip + ':' + nsae_slb_real_port + '##limit#' + nsae_slb_real_limit + "##check#" + nsae_slb_real_check + '##'

    nsae_slb_real_disable_map = {}
    nsae_slb_real_disable_list = nsae_slb_real_disable_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_real_disable_list:
        nsae_slb_real_name = re.sub('\s*','',item)
        nsae_slb_real_disable_map[nsae_slb_real_name] = "disable"

    nsae_slb_group_member_map = {}
    nsae_slb_group_member_list = nsae_slb_group_member_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_group_member_list:
        nsae_slb_pool_name = re.sub('\s*','',item[0])
        nsae_slb_pool_member = re.sub('\s*','',item[1])
        nsae_slb_pool_info = nsae_slb_pool_member + '\n'
        if nsae_slb_pool_name in nsae_slb_group_member_map.keys():
            nsae_slb_pool_info = nsae_slb_pool_info + nsae_slb_group_member_map[nsae_slb_pool_name]

        nsae_slb_group_member_map[nsae_slb_pool_name] = nsae_slb_pool_info

    nsae_slb_virtual_map = {}
    nsae_slb_virtual_list = nsae_slb_virtual_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_virtual_list:
        nsae_slb_vs_name = re.sub('\s*','',item[0])
        nsae_slb_vs_ip = re.sub('\s*','',item[1])
        nsae_slb_vs_port = re.sub('\s*','',item[2])
        nsae_slb_vs_info = nsae_slb_vs_ip + ':' + nsae_slb_vs_port
        nsae_slb_virtual_map[nsae_slb_vs_name] = nsae_slb_vs_info

    nsae_slb_policy_map = {}
    nsae_slb_policy_list = nsae_slb_policy_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_policy_list:
        nsae_slb_vs_name = re.sub('\s*','',item[0])
        nsae_slb_vs_pool = re.sub('\s*','',item[1])
        nsae_slb_policy_map[nsae_slb_vs_name] = nsae_slb_vs_pool

    nsae_ssl_host_map = {}
    nsae_ssl_host_list = nsae_ssl_host_pattern.findall(ssl_config_open_str)
    for item in nsae_ssl_host_list:
        nsae_ssl_host_name = re.sub('\s*','',item[0])
        nsae_slb_vs_name = re.sub('\s*','',item[1])
        nsae_ssl_host_map[nsae_slb_vs_name] = nsae_ssl_host_name

    nsae_ssl_vs_list = []
    for vs_name in nsae_slb_virtual_map.keys():
        nsae_ssl_vs_info = ['']*6
        nsae_ssl_vs_info[0] = nsae_slb_virtual_map[vs_name]
        nsae_ssl_vs_info[1] = vs_name
        nsae_ssl_vs_info[2] = ''
        if vs_name in nsae_ssl_host_map.keys():
            nsae_ssl_vs_info[2] = nsae_ssl_host_map[vs_name]
        nsae_ssl_vs_pool_name = ''
        nsae_ssl_vs_member_simple = ''
        nsae_ssl_vs_member_detail = ''
        if vs_name in nsae_slb_policy_map.keys():
            nsae_ssl_vs_pool_name = nsae_slb_policy_map[vs_name].strip()
            if nsae_ssl_vs_pool_name in nsae_slb_group_member_map.keys():
                real_members = nsae_slb_group_member_map[nsae_ssl_vs_pool_name].strip()
                real_members_list = real_members.split('\n')
                vs_member_detail_info = ''
                vs_member_simple_info = ''
                for real_member in real_members_list:
                    if real_member != '' and real_member is not None:
                        real_member_info = nsae_slb_real_map[real_member]
                        real_member_ipport_pattern = re.compile("##ipport#([\s\S]*?)##", re.MULTILINE)
                        real_member_ipport = ''.join(real_member_ipport_pattern.findall(real_member_info))
                        real_member_limit_pattern = re.compile("##limit#([\s\S]*?)##", re.MULTILINE)
                        real_member_limit = ''.join(real_member_limit_pattern.findall(real_member_info))
                        real_member_check_pattern = re.compile("##check#([\s\S]*?)##", re.MULTILINE)
                        real_member_check = ''.join(real_member_check_pattern.findall(real_member_info))
                        if real_member in nsae_slb_real_disable_map.keys():
                            vs_member_detail_info += real_member_ipport + ' disable' + ' l:' + real_member_limit + ' c:' + real_member_check + '\n'
                        else:
                            vs_member_detail_info += real_member_ipport + ' enable' + ' l:' + real_member_limit + ' c:' + real_member_check + '\n'
                            vs_member_simple_info += real_member_ipport + '\n'

                nsae_ssl_vs_member_simple = vs_member_simple_info.strip('\n')
                nsae_ssl_vs_member_detail = vs_member_detail_info.strip('\n')
            elif nsae_ssl_vs_pool_name in nsae_slb_real_map.keys():
                real_member_info = nsae_slb_real_map[nsae_ssl_vs_pool_name]
                real_member_ipport_pattern = re.compile("##ipport#([\s\S]*?)##", re.MULTILINE)
                real_member_ipport = ''.join(real_member_ipport_pattern.findall(real_member_info))
                real_member_limit_pattern = re.compile("##limit#([\s\S]*?)##", re.MULTILINE)
                real_member_limit = ''.join(real_member_limit_pattern.findall(real_member_info))
                real_member_check_pattern = re.compile("##check#([\s\S]*?)##", re.MULTILINE)
                real_member_check = ''.join(real_member_check_pattern.findall(real_member_info))
                vs_member_detail_info = real_member_ipport + ' enable' + ' l:' + real_member_limit + ' c:' + real_member_check
                vs_member_simple_info = ''
                if nsae_ssl_vs_pool_name in nsae_slb_real_disable_map.keys():
                    vs_member_detail_info = real_member_ipport + ' disable' + ' l:' + real_member_limit + ' c:' + real_member_check
                else:
                    vs_member_simple_info = vs_member_simple_info + real_member_ipport

                nsae_ssl_vs_member_simple = vs_member_simple_info
                nsae_ssl_vs_member_detail = vs_member_detail_info

        nsae_ssl_vs_info[3] = nsae_ssl_vs_pool_name
        nsae_ssl_vs_info[4] = nsae_ssl_vs_member_simple
        nsae_ssl_vs_info[5] = nsae_ssl_vs_member_detail
        nsae_ssl_vs_list.append(nsae_ssl_vs_info)

    ssl_config_open.close()
    return nsae_ssl_vs_list



def main():
    nas_fliename_list = get_nas_filename_list()
    ltm_device_list = get_device_data(analyzer_path)

    networks = {}
    for device in ltm_device_list:
        if device not in nas_fliename_list.keys():
            print(device+"设备名不正确，请填入正确的设备名称！")
            break
        filepath = config_path + '\\' + nas_fliename_list[device]
        device_net_info = [''] * 6
        version = NAS_DEVICE_DIR[device]['version']
        real_name = NAS_DEVICE_DIR[device]['real_name']
        mgmt_ip = NAS_DEVICE_DIR[device]['mgmt_ip']
        device_type = NAS_DEVICE_DIR[device]['type']
        # ltm_config = get_ltm_base_config(filepath,device_type,version)
        # device_net_info[0] = real_name
        # device_net_info[1] = mgmt_ip
        # device_net_info[2] = ltm_config['self_ip']
        # device_net_info[3] = ltm_config['float_ip']
        # device_net_info[4] = ltm_config['route']
        # device_net_info[5] = ltm_config['acls']
        # networks[device] = device_net_info

    now_time = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    # result_all_networks_list = networks.values()
    # df = pd.DataFrame(result_all_networks_list, columns=['设备名称','管理ip','互联ip','浮动ip','路由','访问控制'])

    # respath = result_path + "result_all_networks_" + now_time + ".xlsx"
    # df.to_excel(respath, index=False)
    # print('解析完成：'+respath)
    #
    # ssl_vs_info_lsit = get_ssl_config(filepath, device_type, version)
    #
    # df2 = pd.DataFrame(ssl_vs_info_lsit, columns=['vs名称', 'vs的ip和端口', 'ssl_profile名称','域名','证书过期时间'])
    # respath2 = result_path + "result_all_ssl_" + now_time + ".xlsx"
    #
    # df2.to_excel(respath2, index=False)
    #
    # print('解析完成：'+respath2)
    #
    # ssl_exp_info_map = get_ssl_exp_config(filepath, device_type, version)
    #
    # ssl_cert_exp_info_lsit = ssl_exp_info_map['ssl_exp_not_del_info']
    #
    # df3 = pd.DataFrame(ssl_cert_exp_info_lsit, columns=['vs名称','ssl_profile名称', '证书名称', '私钥名称', 'chain证书名称','CA证书名称','域名','证书过期时间'])
    # respath3 = result_path + "result_all_exp_ssl_cert_not_del_" + now_time + ".xlsx"
    #
    # df3.to_excel(respath3, index=False)
    #
    # print('解析完成：'+respath3)
    #
    # ssl_cert_exp_can_del_lsit = ssl_exp_info_map['ssl_exp_can_del_info']
    #
    # df4 = pd.DataFrame(ssl_cert_exp_can_del_lsit, columns=['ssl_profile名称', '证书名称', '私钥名称', 'chain证书名称','CA证书名称','域名','证书过期时间'])
    # respath4 = result_path + "result_all_exp_ssl_cert_can_del_" + now_time + ".xlsx"
    #
    # df4.to_excel(respath4, index=False)
    #
    # print('解析完成：'+respath4)
    #
    #
    # ltm_v12_ssl_exp_cert_list = ssl_exp_info_map['ssl_exp_cert_info']
    #
    # df5 = pd.DataFrame(ltm_v12_ssl_exp_cert_list, columns=['证书名称','域名','证书过期时间'])
    # respath5 = result_path + "result_all_exp_ssl_cert_info_" + now_time + ".xlsx"
    #
    # df5.to_excel(respath5, index=False)
    #
    # print('解析完成：'+respath5)
    #
    #
    # ltm_v12_vs_list =  get_ltm_config(filepath, type, version)
    #
    # df6 = pd.DataFrame(ltm_v12_vs_list, columns=['vs名称', 'vs连接数限制', 'vs的ip_port', 'vs_status', 'vs_protocol', 'vs_persist_name', 'vs_persist_mothod', 'vs_persist_timeout', 'persist_cookie_encrypt', 'persist_cookie_name', 'persist_cookie_method', 'vs_pool_name', 'vs_balanc_mode', 'vs_pool_monitor', 'members_info_simple', 'members_info_detail', 'fastl4_profile_name', 'fastl4_timeout', 'fastl4_pva', 'tcp_profile_name', 'tcp_profile_timeout', 'http_profile_name', 'http_profile_xforwarded', 'other_profile', 'vs_rules', 'vs_snat_pool_name', 'vs_source_port', 'vs_vlans'])
    # respath6 = result_path + "result_all_vs_info_" + now_time + ".xlsx"
    #
    # df6.to_excel(respath6, index=False)
    #
    # print('解析完成：'+respath6)


    nsae_ssl_vs_list =  get_nsae_ssl_config(filepath, type, version)

    df7 = pd.DataFrame(nsae_ssl_vs_list, columns=['ssl_vs_ipport', 'ssl_vs_name', 'ssl_host', 'ssl_pool_name', 'members_info_simple', 'members_info_detail'])
    respath7 = result_path + "nsae_ssl_vs_info_" + now_time + ".xlsx"

    df7.to_excel(respath7, index=False)

    print('解析完成：'+respath7)

if __name__ == '__main__':
    main()






