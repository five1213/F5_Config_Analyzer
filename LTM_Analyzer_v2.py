import configparser
import json
import re
import datetime
import os
import pandas as pd
from openpyxl import load_workbook


config_os = 'linux'
config_path = '/Users/device'
#config_os = 'windows'
#config_path = 'E:\\config'


device_analyzer_list = []
device_list_map = {}
device_path_map = {}

device_list_path = ''
config_path_os = ''
if config_os == 'windows':
    config_path_os = config_path + '\\'
elif config_os == 'linux':
    config_path_os = config_path + '/'

device_list_path = config_path_os + '设备列表.xlsx'
waf_pass_path = config_path_os + 'waf旁路表.xlsx'

port_json_str_open = open('config/port.json', encoding='utf-8' ,errors='ignore')
port_json_str = port_json_str_open.read()
ports_dir = json.loads(port_json_str)
ports_data = dict([val, key] for key, val in ports_dir.items())
port_json_str_open.close()

config = configparser.ConfigParser()
config.read('config/config_v2.ini', encoding='utf-8')

LTM_V12_SOURCE_PERSIST_RE_STR = config.get('LTM', 'LTM_V12_SOURCE_PERSIST_RE_STR')
LTM_V12_COOKIE_PERSIST_RE_STR = config.get('LTM', 'LTM_V12_COOKIE_PERSIST_RE_STR')
LTM_V12_HTTP_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_HTTP_PROFILE_RE_STR')
LTM_V12_TCP_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_TCP_PROFILE_RE_STR')
LTM_V12_FASTL4_PROFILE_RE_STR = config.get('LTM', 'LTM_V12_FASTL4_PROFILE_RE_STR')
LTM_V12_POOL_RE_STR = config.get('LTM', 'LTM_V12_POOL_RE_STR')
LTM_V12_POOL_MEMBER_RE_STR = config.get('LTM', 'LTM_V12_POOL_MEMBER_RE_STR')
LTM_V12_VS_RE_STR = config.get('LTM', 'LTM_V12_VS_RE_STR')

NSAE_SLB_REAL_RE_STR = config.get('LTM', 'NSAE_SLB_REAL_RE_STR')
NSAE_SLB_REAL_DISABLE_RE_STR = config.get('LTM', 'NSAE_SLB_REAL_DISABLE_RE_STR')
NSAE_SLB_GROUP_MEMBER_RE_STR = config.get('LTM', 'NSAE_SLB_GROUP_MEMBER_RE_STR')
NSAE_SLB_VIRTUAL_RE_STR = config.get('LTM', 'NSAE_SLB_VIRTUAL_RE_STR')
NSAE_SLB_POLICY_RE_STR = config.get('LTM', 'NSAE_SLB_POLICY_RE_STR')
NSAE_SSL_HOST_RE_STR = config.get('LTM', 'NSAE_SSL_HOST_RE_STR')

CITRIX_POOl_MEM_RE_STR = config.get('LTM', 'CITRIX_POOl_MEM_RE_STR')
CITRIX_POOl_RE_STR = config.get('LTM', 'CITRIX_POOl_RE_STR')
CITRIX_VS_POLICY_RE_STR = config.get('LTM', 'CITRIX_VS_POLICY_RE_STR')
CITRIX_VS_RE_STR = config.get('LTM', 'CITRIX_VS_RE_STR')
CITRIX_POLICY_RE_STR = config.get('LTM', 'CITRIX_POLICY_RE_STR')
CITRIX_SERVER_RE_STR = config.get('LTM', 'CITRIX_SERVER_RE_STR')

LTM_V12_ROUTE_RE_STR = config.get('LTM', 'LTM_V12_ROUTE_RE_STR')
LTM_V12_MGT_ROUTE_RE_STR = config.get('LTM', 'LTM_V12_MGT_ROUTE_RE_STR')
LTM_V12_SELF_IP_RE_STR = config.get('LTM', 'LTM_V12_SELF_IP_RE_STR')
LTM_V10_ROUTE_RE_STR = config.get('LTM', 'LTM_V10_ROUTE_RE_STR')
LTM_V10_MGT_ROUTE_RE_STR = config.get('LTM', 'LTM_V10_MGT_ROUTE_RE_STR')
LTM_V10_SELF_IP_RE_STR = config.get('LTM', 'LTM_V10_SELF_IP_RE_STR')
LTM_HTTP_ACL_RE_STR = config.get('LTM', 'LTM_HTTP_ACL_RE_STR')
LTM_SSH_ACL_RE_STR = config.get('LTM', 'LTM_SSH_ACL_RE_STR')
LTM_NTP_RE_STR = config.get('LTM', 'LTM_NTP_RE_STR')
LTM_SNMP_RE_STR = config.get('LTM', 'LTM_SNMP_RE_STR')
LTM_SYSLOG_RE_STR = config.get('LTM', 'LTM_SYSLOG_RE_STR')
LTM_SNATPOOL_RE_STR = config.get('LTM', 'LTM_SNATPOOL_RE_STR')
NSAE_IP_RE_STR = config.get('LTM', 'NSAE_IP_RE_STR')
NSAE_ROUTE_RE_STR = config.get('LTM', 'NSAE_ROUTE_RE_STR')
NSAE_HTTP_ACL_RE_STR = config.get('LTM', 'NSAE_HTTP_ACL_RE_STR')
NSAE_SSH_ACL_RE_STR = config.get('LTM', 'NSAE_SSH_ACL_RE_STR')
NSAE_NTP_RE_STR = config.get('LTM', 'NSAE_NTP_RE_STR')
NSAE_SNMP_RE_STR = config.get('LTM', 'NSAE_SNMP_RE_STR')
NSAE_SYSLOG_RE_STR = config.get('LTM', 'NSAE_SYSLOG_RE_STR')
NSAE_SNATPOOL_RE_STR = config.get('LTM', 'NSAE_SNATPOOL_RE_STR')
CITRIX_IP_RE_STR = config.get('LTM', 'CITRIX_IP_RE_STR')
CITRIX_ROUTE_RE_STR = config.get('LTM', 'CITRIX_ROUTE_RE_STR')
CITRIX_ACL_RE_STR = config.get('LTM', 'CITRIX_ACL_RE_STR')
CITRIX_SNMP_RE_STR = config.get('LTM', 'CITRIX_SNMP_RE_STR')
CITRIX_SYSLOG_RE_STR = config.get('LTM', 'CITRIX_SYSLOG_RE_STR')
CITRIX_IPSET_RE_STR = config.get('LTM', 'CITRIX_IPSET_RE_STR')
CITRIX_SNATPOOL_RE_STR = config.get('LTM', 'CITRIX_SNATPOOL_RE_STR')

ltm_v12_source_persist_pattern = re.compile(LTM_V12_SOURCE_PERSIST_RE_STR, re.MULTILINE)
ltm_v12_cookie_persist_pattern = re.compile(LTM_V12_COOKIE_PERSIST_RE_STR, re.MULTILINE)
ltm_v12_http_profile_pattern = re.compile(LTM_V12_HTTP_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_tcp_profile_pattern = re.compile(LTM_V12_TCP_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_fastl4_profile_pattern = re.compile(LTM_V12_FASTL4_PROFILE_RE_STR, re.MULTILINE)
ltm_v12_pool_pattern = re.compile(LTM_V12_POOL_RE_STR, re.MULTILINE)
ltm_v12_pool_member_pattern = re.compile(LTM_V12_POOL_MEMBER_RE_STR, re.MULTILINE)
ltm_v12_vs_pattern = re.compile(LTM_V12_VS_RE_STR, re.MULTILINE)

nsae_slb_real_pattern = re.compile(NSAE_SLB_REAL_RE_STR, re.MULTILINE)
nsae_slb_real_disable_pattern = re.compile(NSAE_SLB_REAL_DISABLE_RE_STR, re.MULTILINE)
nsae_slb_group_member_pattern = re.compile(NSAE_SLB_GROUP_MEMBER_RE_STR, re.MULTILINE)
nsae_slb_virtual_pattern = re.compile(NSAE_SLB_VIRTUAL_RE_STR, re.MULTILINE)
nsae_slb_policy_pattern = re.compile(NSAE_SLB_POLICY_RE_STR, re.MULTILINE)
nsae_ssl_host_pattern = re.compile(NSAE_SSL_HOST_RE_STR, re.MULTILINE)

citrix_pool_mem_pattern = re.compile(CITRIX_POOl_MEM_RE_STR, re.MULTILINE)
citrix_pool_pattern = re.compile(CITRIX_POOl_RE_STR, re.MULTILINE)
citrix_vs_policy_pattern = re.compile(CITRIX_VS_POLICY_RE_STR, re.MULTILINE)
citrix_vs_pattern = re.compile(CITRIX_VS_RE_STR, re.MULTILINE)
citrix_policy_pattern = re.compile(CITRIX_POLICY_RE_STR, re.MULTILINE)
citrix_server_pattern = re.compile(CITRIX_SERVER_RE_STR, re.MULTILINE)

ltm_v12_route_pattern = re.compile(LTM_V12_ROUTE_RE_STR, re.MULTILINE)
ltm_v12_mgt_route_pattern = re.compile(LTM_V12_MGT_ROUTE_RE_STR, re.MULTILINE)
ltm_v12_self_ip_pattern = re.compile(LTM_V12_SELF_IP_RE_STR, re.MULTILINE)
ltm_v10_route_pattern = re.compile(LTM_V10_ROUTE_RE_STR, re.MULTILINE)
ltm_v10_mgt_route_pattern = re.compile(LTM_V10_MGT_ROUTE_RE_STR, re.MULTILINE)
ltm_v10_self_ip_pattern = re.compile(LTM_V10_SELF_IP_RE_STR, re.MULTILINE)
ltm_http_acl_pattern = re.compile(LTM_HTTP_ACL_RE_STR, re.MULTILINE)
ltm_ssh_acl_pattern = re.compile(LTM_SSH_ACL_RE_STR, re.MULTILINE)
ltm_ntp_pattern = re.compile(LTM_NTP_RE_STR, re.MULTILINE)
ltm_snmp_pattern = re.compile(LTM_SNMP_RE_STR, re.MULTILINE)
ltm_syslog_pattern = re.compile(LTM_SYSLOG_RE_STR, re.MULTILINE)
ltm_snatpool_pattern = re.compile(LTM_SNATPOOL_RE_STR, re.MULTILINE)
nsae_ip_pattern = re.compile(NSAE_IP_RE_STR, re.MULTILINE)
nsae_route_pattern = re.compile(NSAE_ROUTE_RE_STR, re.MULTILINE)
nsae_http_acl_pattern = re.compile(NSAE_HTTP_ACL_RE_STR, re.MULTILINE)
nsae_ssh_acl_pattern = re.compile(NSAE_SSH_ACL_RE_STR, re.MULTILINE)
nsae_ntp_pattern = re.compile(NSAE_NTP_RE_STR, re.MULTILINE)
nsae_snmp_pattern = re.compile(NSAE_SNMP_RE_STR, re.MULTILINE)
nsae_syslog_pattern = re.compile(NSAE_SYSLOG_RE_STR, re.MULTILINE)
nsae_snatpool_pattern = re.compile(NSAE_SNATPOOL_RE_STR, re.MULTILINE)
citrix_ip_pattern = re.compile(CITRIX_IP_RE_STR, re.MULTILINE)
citrix_route_pattern = re.compile(CITRIX_ROUTE_RE_STR, re.MULTILINE)
citrix_acl_pattern = re.compile(CITRIX_ACL_RE_STR, re.MULTILINE)
citrix_snmp_pattern = re.compile(CITRIX_SNMP_RE_STR, re.MULTILINE)
citrix_syslog_pattern = re.compile(CITRIX_SYSLOG_RE_STR, re.MULTILINE)
citrix_ipset_pattern = re.compile(CITRIX_IPSET_RE_STR, re.MULTILINE)
citrix_snatpool_pattern = re.compile(CITRIX_SNATPOOL_RE_STR, re.MULTILINE)

def get_device_list():
    wb = load_workbook(device_list_path)  # 打开Excel
    sheet1 = wb['解析列表']
    for row in range(2, sheet1.max_row + 1):
        device_analyzer = sheet1.cell(row, 1).value.strip()
        if device_analyzer == '':
            print('解析列表为空，请填写要解析的设备！')
            return
        device_analyzer_list.append(device_analyzer)

    sheet2 = wb['设备列表']
    for row in range(2, sheet2.max_row + 1):
        name_device = sheet2.cell(row, 1).value.strip()
        if name_device != '':
            device_info = {}
            device_info['name'] = name_device
            device_info['type'] = sheet2.cell(row, 2).value.strip()
            device_info['version'] = sheet2.cell(row, 3).value.strip()
            device_list_map[name_device] = device_info

    files= os.listdir(config_path)
    for file in files:
        new_file_path = config_path_os + file
        if os.path.isdir(new_file_path):
            devices = os.listdir(new_file_path)
            for device in devices:
                devicename,extension = os.path.splitext(device)
                if extension == '.txt' or extension == '.TXT':
                    if config_os == 'windows':
                        device_path_map[devicename] = new_file_path +'\\'+ device
                    elif config_os == 'linux':
                        device_path_map[devicename] = new_file_path + '/' + device
waf_paas_map = {}
waf_over_map = {}
ssl_update_map = {}
scripts_paas_map = {}
scripts_over_map = {}
def get_waf_pass_list():
    wb = load_workbook(waf_pass_path)  # 打开Excel
    sheet1 = wb['waf']
    for row in range(2, sheet1.max_row + 1):
        sys_name = sheet1.cell(row, 1).value.strip()
        new_sys_name = sys_name
        if 'v4' in sys_name or 'V4' in sys_name:
            new_sys_name = sys_name.replace('4','6')

        domain = sheet1.cell(row, 2).value.strip()

        ssl_vs = sheet1.cell(row, 3).value.strip()
        if not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$",ssl_vs):
            ssl_vs = re.sub(':0{1,}',':',ssl_vs.lower())

        ssl_members = sheet1.cell(row, 4).value

        waf_vs = sheet1.cell(row, 5).value
        if waf_vs != None and not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$",waf_vs):
            waf_vs = re.sub(':0{1,}',':',waf_vs.lower())

        web_vs = sheet1.cell(row, 6).value.strip()
        if not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$",web_vs):
            web_vs = re.sub(':0{1,}',':',web_vs.lower())

        if ssl_members != None:
            members = ssl_members.strip().split("\n")
            for nsae_vs in members:
                if not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$",nsae_vs):
                    nsae_vs = re.sub(':0{1,}',':',nsae_vs.lower())

                if nsae_vs in nsae_vs_device_map.keys():
                    device_name = nsae_vs_device_map[nsae_vs]
                    nsae_waf_vs_mem = nsae_real_map[device_name + '_' + waf_vs]
                    nsae_web_vs_mem = nsae_real_map[device_name + '_' + web_vs]
                    device_name_pattern = re.compile("[\s\S]*?(lb[\d]*|ssl[\d]*)", re.MULTILINE)
                    pre_str = ''.join(device_name_pattern.findall(device_name))
                    if re.match(r"^(ssl[\d])$", pre_str):
                        pre_str = re.sub('ssl', 'ssl0', pre_str)
                    if re.match(r"^(lb[\d])$", pre_str):
                        pre_str = re.sub('lb', 'lb0', pre_str)

                    waf_paas_enable_script = pre_str + '##'+ sys_name + '##' + device_name + '##' + '0' + '##' + 'slb real enable "' + nsae_web_vs_mem + '"' +'\n'
                    pass_enable_key = domain+'_' + waf_paas_enable_script
                    if pass_enable_key not in scripts_paas_map.keys():
                        scripts_paas_map[pass_enable_key] = ''
                        if domain in waf_paas_map.keys():
                            waf_paas_map[domain] = waf_paas_map[domain] + waf_paas_enable_script
                        else:
                            waf_paas_map[domain] = waf_paas_enable_script

                    waf_paas_disable_script = pre_str + '##' + sys_name + '##' + device_name + '##' + '1' + '##' + 'slb real disable "' + nsae_waf_vs_mem + '"' + '\n'
                    pass_disable_key = domain + '_' + waf_paas_disable_script
                    if pass_disable_key not in scripts_paas_map.keys():
                        scripts_paas_map[pass_disable_key] = ''
                        if domain in waf_paas_map.keys():
                            waf_paas_map[domain] = waf_paas_map[domain] + waf_paas_disable_script
                        else:
                            waf_paas_map[domain] = waf_paas_disable_script

                    waf_over_enable_script = pre_str + '##'+ sys_name + '##' + device_name + '##' + '0' + '##' + 'slb real enable "' + nsae_waf_vs_mem + '"'+ '\n'
                    over_enable_key = domain + '_' + waf_over_enable_script
                    if over_enable_key not in scripts_over_map.keys():
                        scripts_over_map[over_enable_key] = ''
                        if domain in waf_over_map.keys():
                            waf_over_map[domain] = waf_over_map[domain] + waf_over_enable_script
                        else:
                            waf_over_map[domain] = waf_over_enable_script

                    waf_over_disable_script = pre_str + '##' + sys_name + '##' + device_name + '##' + '1' + '##' + 'slb real disable "' + nsae_web_vs_mem + '"' + '\n'
                    over_disable_key = domain + '_' + waf_over_disable_script
                    if over_disable_key not in scripts_over_map.keys():
                        scripts_over_map[over_disable_key] = ''
                        if domain in waf_over_map.keys():
                            waf_over_map[domain] = waf_over_map[domain] + waf_over_disable_script
                        else:
                            waf_over_map[domain] = waf_over_disable_script

                    ssl_host_info = nsae_vs_sslhost_map[nsae_vs]
                    ssl_update_script = pre_str + '##'+ sys_name + '##' + ssl_host_info + '\n'
                    if domain in ssl_update_map.keys():
                        ssl_update_map[domain] = ssl_update_map[domain] + ssl_update_script
                    else:
                        ssl_update_map[domain] = ssl_update_script

        else:
            new_device_name = 'nh7402b0539-nint-lb11'
            if re.match(r"^(10.6.20.[\s\S]*?)$",ssl_vs) or re.match(r"^(2404:bc0:3:114[\s\S]*?)$",ssl_vs):
                new_device_name = 'nh7402b0636-nint-lb21'

            device_name_pattern = re.compile("[\s\S]*?(lb[\d]*|ssl[\d]*)", re.MULTILINE)
            pre_str = ''.join(device_name_pattern.findall(new_device_name))
            if re.match(r"^(ssl[\d])$", pre_str) :
                pre_str = re.sub('ssl', 'ssl0', pre_str)
            if re.match(r"^(lb[\d])$", pre_str) :
                pre_str = re.sub('lb', 'lb0', pre_str)

            ssl_vs_info = f5_vs_info_map[ssl_vs]
            ssl_vs_name_pattern = re.compile("##vs_name#([\s\S]*?)##", re.MULTILINE)
            ssl_vs_name = ''.join(ssl_vs_name_pattern.findall(ssl_vs_info))

            ssl_vs_persist_name_pattern = re.compile("##vs_persist_name#([\s\S]*?)##", re.MULTILINE)
            ssl_vs_persist_name = ''.join(ssl_vs_persist_name_pattern.findall(ssl_vs_info))

            ssl_vs_pool_name_pattern = re.compile("##vs_pool_name#([\s\S]*?)##", re.MULTILINE)
            ssl_vs_pool_name = ''.join(ssl_vs_pool_name_pattern.findall(ssl_vs_info))

            ssl_vs_profile_name_pattern = re.compile("##vs_profile_name#([\s\S]*?)##", re.MULTILINE)
            ssl_vs_profile_name = ''.join(ssl_vs_profile_name_pattern.findall(ssl_vs_info))

            ssl_vs_snat_pool_name_pattern = re.compile("##vs_snat_pool_name#([\s\S]*?)##", re.MULTILINE)
            ssl_vs_snat_pool_name = ''.join(ssl_vs_snat_pool_name_pattern.findall(ssl_vs_info))

            web_vs_info = f5_vs_info_map[web_vs]
            web_vs_name_pattern = re.compile("##vs_name#([\s\S]*?)##", re.MULTILINE)
            web_vs_name = ''.join(web_vs_name_pattern.findall(web_vs_info))

            web_vs_persist_name_pattern = re.compile("##vs_persist_name#([\s\S]*?)##", re.MULTILINE)
            web_vs_persist_name = ''.join(web_vs_persist_name_pattern.findall(web_vs_info))

            web_vs_pool_name_pattern = re.compile("##vs_pool_name#([\s\S]*?)##", re.MULTILINE)
            web_vs_pool_name = ''.join(web_vs_pool_name_pattern.findall(web_vs_info))

            web_vs_profile_name_pattern = re.compile("##vs_profile_name#([\s\S]*?)##", re.MULTILINE)
            web_vs_profile_name = ''.join(web_vs_profile_name_pattern.findall(web_vs_info))

            web_vs_snat_pool_name_pattern = re.compile("##vs_snat_pool_name#([\s\S]*?)##", re.MULTILINE)
            web_vs_snat_pool_name = ''.join(web_vs_snat_pool_name_pattern.findall(web_vs_info))

            waf_paas_script =  pre_str + '##'+sys_name + '##' + new_device_name + '##' + '0' + '##' + 'tmsh modify ltm virtual ' + ssl_vs_name  + ' profiles replace-all-with { ' + web_vs_profile_name + ' } ' + 'persist replace-all-with { ' + web_vs_persist_name + ' }' + ' pool ' + web_vs_pool_name + ' source-port change snatpool ' + web_vs_snat_pool_name + '\n'
            if web_vs_persist_name == 'none':
                waf_paas_script = pre_str + '##' + sys_name + '##' + new_device_name + '##' + '0' + '##' + 'tmsh modify ltm virtual ' + ssl_vs_name + ' profiles replace-all-with { ' + web_vs_profile_name + ' } ' + 'persist none' + ' pool ' + web_vs_pool_name + ' source-port change snatpool ' + web_vs_snat_pool_name + '\n'
            waf_paas_key = domain + '_' + waf_paas_script
            if waf_paas_key not in scripts_paas_map.keys():
                scripts_paas_map[waf_paas_key] = ''
                if domain in waf_paas_map.keys():
                    waf_paas_map[domain] = waf_paas_map[domain] + waf_paas_script
                else:
                    waf_paas_map[domain] = waf_paas_script

            waf_paas_syn_script = pre_str + '##' + new_sys_name + '##' + new_device_name + '##' + '1' + '##' + 'tmsh run cm config-sync to-group Device-Group' + '\n'
            waf_paas_syn_key = domain + '_' + waf_paas_syn_script
            if waf_paas_syn_key not in scripts_paas_map.keys():
                scripts_paas_map[waf_paas_syn_key] = ''
                if domain in waf_paas_map.keys():
                    waf_paas_map[domain] = waf_paas_map[domain] + waf_paas_syn_script
                else:
                    waf_paas_map[domain] = waf_paas_syn_script

            waf_over_script =  pre_str + '##' + sys_name + '##' + new_device_name + '##' + '0' + '##' + 'tmsh modify ltm virtual ' + ssl_vs_name  + ' profiles replace-all-with { ' + ssl_vs_profile_name + ' } ' + 'persist none' + ' pool ' + ssl_vs_pool_name + ' snat none source-port preserve' + '\n'
            waf_over_key = domain + '_' + waf_over_script
            if waf_over_key not in scripts_over_map.keys():
                scripts_over_map[waf_over_key] = ''
                if domain in waf_over_map.keys():
                    waf_over_map[domain] = waf_over_map[domain] + waf_over_script
                else:
                    waf_over_map[domain] = waf_over_script

            waf_over_syn_script = pre_str + '##' + new_sys_name + '##' + new_device_name + '##' + '1' + '##' + 'tmsh run cm config-sync to-group Device-Group' + '\n'
            waf_over_syn_key = domain + '_' + waf_over_syn_script
            if waf_over_syn_key not in scripts_over_map.keys():
                scripts_over_map[waf_over_syn_key] = ''
                if domain in waf_over_map.keys():
                    waf_over_map[domain] = waf_over_map[domain] + waf_over_syn_script
                else:
                    waf_over_map[domain] = waf_over_syn_script
def get_ltm_base_config(filepath,type,version,device_name):

    ltm_config_open = open(filepath, encoding='utf-8' ,errors='ignore')
    ltm_config_open_str = ltm_config_open.read()
    ltm_config_open.close()

    routes = ''
    bus_routes = ''
    mgt_routes = ''
    self_ips = ''
    acls = ''
    ntp = ''
    snmp = ''
    syslog = ''
    snatpool = ''

    if type == 'ltm' or type == 'gtm':
        if version == 'v11' :
            ltm_v12_route = ltm_v12_route_pattern.findall(ltm_config_open_str)
            for item in ltm_v12_route:
                network = item[1].strip()
                gw = item[0].strip()
                bus_routes = bus_routes + network + ' gw ' + gw + '\n'

            ltm_v12_mgt_route = ltm_v12_mgt_route_pattern.findall(ltm_config_open_str)
            for item in ltm_v12_mgt_route:
                network = item[1].strip()
                gateway = item[0].strip()
                mgt_routes = mgt_routes + network + ' gateway ' + gateway + '\n'

            routes = 'bus_routes:\n' + bus_routes + 'mgt_routes:\n' + mgt_routes.strip('\n')

            ltm_v12_self_ip = ltm_v12_self_ip_pattern.findall(ltm_config_open_str)
            for item in ltm_v12_self_ip:
                self_info = item

                address_pattern = re.compile("\s*address\s([\s\S]*?)\n", re.MULTILINE)
                address = ''.join(address_pattern.findall(self_info))

                traffic_group_pattern = re.compile("\s*traffic-group\s([\s\S]*?)\n", re.MULTILINE)
                traffic_group = ''.join(traffic_group_pattern.findall(self_info))

                vlan_pattern = re.compile("\s*vlan\s([\s\S]*?)\n", re.MULTILINE)
                vlan = ''.join(vlan_pattern.findall(self_info))

                self_ips = self_ips + address + ' ' +traffic_group + ' ' + vlan + '\n'

        elif version == 'v10':
            ltm_v10_route = ltm_v10_route_pattern.findall(ltm_config_open_str)
            for item in ltm_v10_route:
                network = item[0].strip()
                gw = item[1].strip()
                bus_routes = bus_routes + network + ' gw ' + gw + '\n'

            ltm_v10_mgt_route = ltm_v10_mgt_route_pattern.findall(ltm_config_open_str)
            for item in ltm_v10_mgt_route:
                network = item[0].strip()
                gateway = item[1].strip()
                mgt_routes = mgt_routes + network + ' gateway ' + gateway + '\n'

            routes = 'bus_routes:\n' + bus_routes + 'mgt_routes:\n' + mgt_routes.strip('\n')

            ltm_v10_self_ip = ltm_v10_self_ip_pattern.findall(ltm_config_open_str)
            for item in ltm_v10_self_ip:
                address = item[0].strip()
                self_info = item[1]

                floating_pattern = re.compile("\s*floating\s([\s\S]*?)\n", re.MULTILINE)
                floating = ''.join(floating_pattern.findall(self_info))
                if floating == '':
                    floating = 'self'
                else:
                    floating = 'floating'

                vlan_pattern = re.compile("\s*vlan\s([\s\S]*?)\n", re.MULTILINE)
                vlan = ''.join(vlan_pattern.findall(self_info))

                self_ips = self_ips + address + ' ' + floating + ' ' + vlan + '\n'


        ltm_http_acl = ltm_http_acl_pattern.findall(ltm_config_open_str)
        if len(ltm_http_acl) > 0:
            https_acls = ltm_http_acl[0].strip().replace(' ','\n')

        ltm_ssh_acl = ltm_ssh_acl_pattern.findall(ltm_config_open_str)
        if len(ltm_ssh_acl) > 0:
            ssh_acls = ltm_ssh_acl[0].strip().replace(' ','\n')

        acls = 'https_acl:\n' + https_acls + '\nssh_acl:\n' + ssh_acls

        ltm_ntp = ltm_ntp_pattern.findall(ltm_config_open_str)
        if len(ltm_ntp) > 0:
            ntp = ltm_ntp[0].strip().replace(' ','\n')

        ltm_snmp = ltm_snmp_pattern.findall(ltm_config_open_str)
        if len(ltm_snmp) > 0:
            snmp = ltm_snmp[0].strip().replace(' ','\n')

        ltm_syslog = ltm_syslog_pattern.findall(ltm_config_open_str)
        if len(ltm_syslog) > 0:
            syslogs = ltm_syslog[0].strip().split(';')
            syslog = '\n'.join(syslogs)

        ltm_snatpools = ltm_snatpool_pattern.findall(ltm_config_open_str)
        for item in ltm_snatpools:
            snatpool_name = item[0].strip()
            snat_address = item[1]
            snatpool = snatpool + snatpool_name + ':\n' + snat_address + '\n'

    elif type == 'nsae':
        nsae_ip = nsae_ip_pattern.findall(ltm_config_open_str)
        for item in nsae_ip:
            interface = item[0].strip().replace('"','')
            ip = item[1].strip()
            mask = item[2].strip()
            self_ips = self_ips + ip + '/' + mask + ' ' + interface + '\n'

        nsae_route = nsae_route_pattern.findall(ltm_config_open_str)
        for item in nsae_route:
            type = item[0].strip()
            route = item[1].strip()
            if type == 'default':
                routes = routes + 'default' + ' gw ' + route + '\n'
            elif type == 'static':
                patterns = r' +'
                network = re.split(patterns,route)
                routes = routes + network[0] + '/' + network[1] + ' gw ' + network[2] + '\n'

        nsae_http_acl = nsae_http_acl_pattern.findall(ltm_config_open_str)
        http_acls = ''
        for item in nsae_http_acl:
            acl_temp = item.strip().replace(' ','/')
            http_acls = http_acls + acl_temp + '\n'

        nsae_ssh_acl = nsae_ssh_acl_pattern.findall(ltm_config_open_str)
        ssh_acls = ''
        for item in nsae_ssh_acl:
            acl_temp = item.strip().replace(' ','/')
            ssh_acls = ssh_acls + acl_temp + '\n'

        acls = 'https_acl:\n' + http_acls + 'ssh_acl:\n' + ssh_acls

        nsae_ntp = nsae_ntp_pattern.findall(ltm_config_open_str)
        for item in nsae_ntp:
            ntp = ntp + item.strip() + '\n'

        nsae_snmp = nsae_snmp_pattern.findall(ltm_config_open_str)
        for item in nsae_snmp:
            snmp = snmp + item.strip() + '\n'

        nsae_syslog = nsae_syslog_pattern.findall(ltm_config_open_str)
        for item in nsae_syslog:
            syslog = syslog + item.strip() + '\n'

        nsae_snatpool = nsae_snatpool_pattern.findall(ltm_config_open_str)
        for item in nsae_snatpool:
            snatpool =  snatpool + item[0] + ': ' + item[1] + '-' + item[2]  + '\n'

    elif type == 'citrix':
        citrix_ip = citrix_ip_pattern.findall(ltm_config_open_str)
        for item in citrix_ip:
            vlan_id = item[0].strip()
            ip = item[1].strip()
            mask = item[2].strip()
            self_ips = self_ips + ip + '/' + mask + ' vlan ' + vlan_id + '\n'

        citrix_route = citrix_route_pattern.findall(ltm_config_open_str)
        for item in citrix_route:
            network = item[0].strip()
            mask = item[1].strip()
            gw = item[2].strip()
            routes = routes + network + '/' + mask + ' gw ' + gw + '\n'

        citrix_acl = citrix_acl_pattern.findall(ltm_config_open_str)
        for item in citrix_acl:
            acl_allow_src = item[0].strip()
            acl_allow_dest = item[1].strip()
            acls = acls + 'src: ' + acl_allow_src + ' dest: ' + acl_allow_dest + '\n'

        citrix_snmp = citrix_snmp_pattern.findall(ltm_config_open_str)
        for item in citrix_snmp:
            snmp = snmp + item.strip() + '\n'

        citrix_syslog = citrix_syslog_pattern.findall(ltm_config_open_str)
        for item in citrix_syslog:
            syslog = syslog + item.strip() + '\n'

        citrix_ipset_map = {}
        citrix_ipset = citrix_ipset_pattern.findall(ltm_config_open_str)
        for item in citrix_ipset:
            ipset_name = item[0].strip()
            ipset_ip = item[1].strip()
            if ipset_name in citrix_ipset_map.keys():
                old_ip =  citrix_ipset_map[ipset_name]
                citrix_ipset_map[ipset_name] = old_ip + ipset_ip + '\n'
            else:
                citrix_ipset_map[ipset_name] = ipset_ip  + '\n'

        citrix_snatpool = citrix_snatpool_pattern.findall(ltm_config_open_str)
        for item in citrix_snatpool:
            snatpool_name = item[0].strip()
            ipset_name = item[1].strip()
            ipsetstrs = ipset_name.split(' ')
            if len(ipsetstrs) > 0:
                ipset_name = ipsetstrs[0].strip()
            ips = citrix_ipset_map[ipset_name]
            if ipset_name in citrix_ipset_map.keys():
                snatpool = snatpool + snatpool_name + ":\n" + ips

    ltm_base_config = ['']*8
    ltm_base_config[0] = device_name
    ltm_base_config[1] = routes.rstrip('\n')
    ltm_base_config[2] = self_ips.rstrip('\n')
    ltm_base_config[3] = acls.rstrip('\n')
    ltm_base_config[4] = ntp.rstrip('\n')
    ltm_base_config[5] = snmp.rstrip('\n')
    ltm_base_config[6] = syslog.rstrip('\n')
    ltm_base_config[7] = snatpool.rstrip('\n')

    return ltm_base_config


f5_vs_info_map = {}
def get_ltm_config(file_path,type,version,device_name):

    ltm_config_open = open(file_path, encoding='utf-8' ,errors='ignore')
    ltm_config_open_str = ltm_config_open.read()
    ltm_config_open.close()

    ltm_v12_source_persist_map = {}
    ltm_v12_source_persist_map['source_addr'] = '3600'
    ltm_v12_source_persist = ltm_v12_source_persist_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_source_persist:
        name = item[0].strip()
        source_persist_info = item[1]
        source_persist_pattern = re.compile("\s*timeout\s(\d*)", re.MULTILINE)
        time_out = ''.join(source_persist_pattern.findall(source_persist_info))
        if time_out == '':
            time_out = '3600'
        ltm_v12_source_persist_map[name] = time_out

    ltm_v12_cookie_persist_map = {}
    ltm_v12_cookie_persist_map['cookie'] = '##encrypt#disabled##name#null##method#insert##'
    ltm_v12_cookie_persist = ltm_v12_cookie_persist_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_cookie_persist:
        name = item[0].strip()
        cookie_persist_info = item[1]
        is_encrypt_pattern = re.compile("\s*cookie-encryption\s(\w*)", re.MULTILINE)
        is_encrypt = ''.join(is_encrypt_pattern.findall(cookie_persist_info))
        cookie_name_pattern = re.compile("\s*cookie-name\s(\w*)", re.MULTILINE)
        cookie_name = ''.join(cookie_name_pattern.findall(cookie_persist_info))
        method_pattern = re.compile("\s*method\s(\w*)", re.MULTILINE)
        method = ''.join(method_pattern.findall(cookie_persist_info))
        if method == '':
            method = 'insert'
        ltm_v12_cookie_persist_map[name] = '##encrypt#'+is_encrypt+'##name#'+cookie_name+'##method#'+method+'##'

    ltm_v12_http_profile_map = {}
    ltm_v12_http_profile_map['http'] = 'disabled'
    ltm_v12_http_profile = ltm_v12_http_profile_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_http_profile:
        name = item[0].strip()
        http_profile_info = item[1]
        http_profile_pattern = re.compile("\s*insert-xforwarded-for\s(\w*)", re.MULTILINE)
        xforwarded = ''.join(http_profile_pattern.findall(http_profile_info))
        ltm_v12_http_profile_map[name] = xforwarded

    ltm_v12_tcp_profile_map = {}
    ltm_v12_tcp_profile_map['tcp'] = '300'
    ltm_v12_tcp_profile = ltm_v12_tcp_profile_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_tcp_profile:
        name = item[0].strip()
        tcp_profile_info = item[1]
        idle_timeout_pattern = re.compile("\s*idle-timeout\s(\d*)", re.MULTILINE)
        idle_timeout = ''.join(idle_timeout_pattern.findall(tcp_profile_info))
        if idle_timeout == '':
            idle_timeout = '300'
        ltm_v12_tcp_profile_map[name] = idle_timeout

    ltm_v12_fastl4_profile_map = {}
    ltm_v12_fastl4_profile_map['fastL4'] = '##timeout#300##pva#full##'
    ltm_v12_fastl4_profile = ltm_v12_fastl4_profile_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_fastl4_profile:
        name = item[0].strip()
        fastl4_profile_info = item[1]
        idle_timeout_pattern = re.compile("\s*idle-timeout\s(\d*)", re.MULTILINE)
        idle_timeout = ''.join(idle_timeout_pattern.findall(fastl4_profile_info))
        if idle_timeout == '':
            idle_timeout = '300'
        pva_pattern = re.compile("\s*pva-acceleration\s(\w*)", re.MULTILINE)
        pva = ''.join(pva_pattern.findall(fastl4_profile_info))
        if pva == '':
            pva = 'full'
        ltm_v12_fastl4_profile_map[name] = '##timeout#'+idle_timeout+'##pva#'+pva+'##'

    ltm_v12_pool_map = {}
    ltm_v12_pool = ltm_v12_pool_pattern.findall(ltm_config_open_str)
    for item in ltm_v12_pool:
        name = item[0].strip()
        pool_info = item[1]
        balanc_mode_pattern = re.compile("\s*load-balancing-mode\s([\s\S]*?)\n", re.MULTILINE)
        balanc_mode = ''.join(balanc_mode_pattern.findall(pool_info))
        if balanc_mode == '':
            balanc_mode = 'round-robin'

        monitor_pattern = re.compile("\s*monitor\s([\s\S]*?)\n", re.MULTILINE)
        monitor = ''.join(monitor_pattern.findall(pool_info))

        members_str_pattern = re.compile("\s*members\s(none|{[\s\S]*?}\s*})", re.MULTILINE)
        members_str = ''.join(members_str_pattern.findall(pool_info))

        members_info = 'none'
        if members_str != 'none' and  members_str != '':
            ltm_v12_pool_member = ltm_v12_pool_member_pattern.findall(members_str)
            members_info_detail = ''
            members_info_simple = ''
            for item2 in ltm_v12_pool_member:
                ip_port_str = item2[0].strip()
                ip_port_info = ''
                if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$",ip_port_str):
                    ipports = ip_port_str.split(":")
                    ip = ipports[0]
                    port = ipports[1]
                    if port in ports_data.keys():
                        port = ports_data[port]
                    ip_port_info = ip + ":" + port
                else:
                    ipports = ip_port_str.split(".")
                    ip = ipports[0]
                    port = ipports[1]
                    if port in ports_data.keys():
                        port = ports_data[port]
                    ip_port_info = ip + "." + port

                info_member = item2[1].strip()

                session_pattern = re.compile("\s*session\s([\s\S]*?)\n", re.MULTILINE)
                session = ''.join(session_pattern.findall(info_member))

                state_pattern = re.compile("\s*state\s([\s\S]*?)\n", re.MULTILINE)
                state = ''.join(state_pattern.findall(info_member))

                con_limit_pattern = re.compile("\s*connection-limit\s(\d*)", re.MULTILINE)
                con_limit = ''.join(con_limit_pattern.findall(info_member))
                if con_limit == '':
                    con_limit = '0'

                priority_pattern = re.compile("\s*priority-group\s(\d*)", re.MULTILINE)
                priority = ''.join(priority_pattern.findall(info_member))
                if priority == '':
                    priority = '0'

                ratio_pattern = re.compile("\s*ratio\s(\d*)", re.MULTILINE)
                ratio = ''.join(ratio_pattern.findall(info_member))
                if ratio == '':
                    ratio = '0'

                members_info_detail = members_info_detail + ip_port_info + ' ' + session + ' ' + state + ' l:' + con_limit + ' p:' + priority + ' r:' + ratio  + '\n'
                if session == 'user-enabled' or session == 'monitor-enabled':
                    members_info_simple  = members_info_simple + ip_port_info + '\n'

                members_info = '##members_info_simple#' + members_info_simple + '##members_info_detail#' + members_info_detail + '##'

        ltm_v12_pool_map[name] = '##balanc_mode#'+balanc_mode+'##monitor#'+monitor + members_info

    ltm_v12_vs_list = []
    ltm_v12_vs = ltm_v12_vs_pattern.findall(ltm_config_open_str)

    for item in ltm_v12_vs:
        vs = ['']*28
        name = item[0].strip()
        vs[0] = name
        vs_info = item[1]

        vs_conn_limit_pattern = re.compile("\s*connection-limit\s(\d*)", re.MULTILINE)
        vs_conn_limit = ''.join(vs_conn_limit_pattern.findall(vs_info))
        if vs_conn_limit == '':
            vs_conn_limit = '0'
        vs[1] = vs_conn_limit

        vs_ip_port_str_pattern = re.compile("\s*destination\s([\s\S]*?)\n", re.MULTILINE)
        vs_ip_port_str = ''.join(vs_ip_port_str_pattern.findall(vs_info))

        vs_ip_port_info = ''
        vs_ip_port_info_new = ''
        if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$", vs_ip_port_str):
            ipports = vs_ip_port_str.split(":")
            ip = ipports[0]
            port = ipports[1]
            if port in ports_data.keys():
                port = ports_data[port]
            vs_ip_port_info = ip + ":" + port
            vs_ip_port_info_new = ip + ":" + port
        elif vs_ip_port_str == 'any:any':
            vs_ip_port_info = '0.0.0.0:0'
            vs_ip_port_info_new = '0.0.0.0:0'
        else:
            ipports = vs_ip_port_str.split(".")
            ip = ipports[0]
            port = ipports[1]
            if port in ports_data.keys():
                port = ports_data[port]
            vs_ip_port_info = ip + "." + port
            vs_ip_port_info_new = re.sub(':0{1,}',':',ip.lower()) + ":" + port

        vs[2] = vs_ip_port_info

        vs_status_pattern = re.compile("\s*[^\d\w]\s(disabled)\n", re.MULTILINE)
        vs_status = ''.join(vs_status_pattern.findall(vs_info))
        if vs_status.strip() == '':
            vs_status = 'enabled'
        vs[3] = vs_status

        vs_protocol_pattern = re.compile("\s*ip-protocol\s([\s\S]*?)\n", re.MULTILINE)
        vs_protocol = ''.join(vs_protocol_pattern.findall(vs_info))
        if vs_protocol == '':
            vs_protocol = 'any'
        vs[4] = vs_protocol

        vs_persist_str_pattern = re.compile("\s*persist\s(none|{[\s\S]*?})\n", re.MULTILINE)
        vs_persist_str = ''.join(vs_persist_str_pattern.findall(vs_info))

        vs_persist_name = 'none'
        vs_persist_mothod = ''
        vs_persist_timeout = ''
        persist_cookie_encrypt = ''
        persist_cookie_name = ''
        persist_cookie_method = ''
        if vs_persist_str != 'none' and vs_persist_str != '':
            vs_persist_str_pattern = re.compile("{\s*([\s\S]*?)\s{\n", re.MULTILINE)
            vs_persist_name = ''.join(vs_persist_str_pattern.findall(vs_persist_str))
            if vs_persist_name in ltm_v12_source_persist_map.keys():
                vs_persist_mothod = 'source_addr'
                vs_persist_timeout = ltm_v12_source_persist_map[vs_persist_name]
            elif vs_persist_name in ltm_v12_cookie_persist_map.keys():
                vs_persist_mothod = 'session_cookie'
                cookie_persist_str = ltm_v12_cookie_persist_map[vs_persist_name]
                cookie_encrypt_pattern = re.compile("##encrypt#([\s\S]*?)##", re.MULTILINE)
                persist_cookie_encrypt = ''.join(cookie_encrypt_pattern.findall(cookie_persist_str))
                cookie_name_pattern = re.compile("##name#([\s\S]*?)##", re.MULTILINE)
                persist_cookie_name = ''.join(cookie_name_pattern.findall(cookie_persist_str))
                cookie_method_pattern = re.compile("##method#([\s\S]*?)##", re.MULTILINE)
                persist_cookie_method = ''.join(cookie_method_pattern.findall(cookie_persist_str))

        vs[5] = vs_persist_name
        vs[6] = vs_persist_mothod
        vs[7] = vs_persist_timeout
        vs[8] = persist_cookie_encrypt
        vs[9] = persist_cookie_name
        vs[10] = persist_cookie_method

        vs_pool_pattern = re.compile("[^{]\n\s*pool\s([\s\S]*?)\n", re.MULTILINE)
        vs_pool = ''.join(vs_pool_pattern.findall(vs_info))

        vs_pool_name = 'none'
        vs_balanc_mode = ''
        vs_pool_monitor = ''
        members_info_simple = ''
        members_info_detail = ''
        if vs_pool != 'none' and vs_pool != '':
            vs_pool_name = vs_pool
            if vs_pool_name in ltm_v12_pool_map.keys():
                vs_pool_info_str = ltm_v12_pool_map[vs_pool_name]
                vs_balanc_mode_pattern = re.compile("##balanc_mode#([\s\S]*?)##", re.MULTILINE)
                vs_balanc_mode = ''.join(vs_balanc_mode_pattern.findall(vs_pool_info_str))
                vs_pool_monitor_pattern = re.compile("##monitor#([\s\S]*?)##", re.MULTILINE)
                vs_pool_monitor = ''.join(vs_pool_monitor_pattern.findall(vs_pool_info_str))
                members_info_simple_pattern = re.compile("##members_info_simple#([\s\S]*?)##", re.MULTILINE)
                members_info_simple = ''.join(members_info_simple_pattern.findall(vs_pool_info_str))
                members_info_detail_pattern = re.compile("##members_info_detail#([\s\S]*?)##", re.MULTILINE)
                members_info_detail = ''.join(members_info_detail_pattern.findall(vs_pool_info_str))

        vs[11] = vs_pool_name
        vs[12] = vs_balanc_mode
        vs[13] = vs_pool_monitor
        vs[14] = members_info_simple.strip('\n')
        vs[15] = members_info_detail.strip('\n')

        vs_profiles_pattern = re.compile("\s*profiles\s{([\s\S]*?{\s}\n)\s*}\n", re.MULTILINE)
        vs_profiles = ''.join(vs_profiles_pattern.findall(vs_info))

        fastl4_profile_name = ''
        fastl4_timeout = ''
        fastl4_pva = ''
        tcp_profile_name = ''
        tcp_profile_timeout = ''
        http_profile_name = ''
        http_profile_xforwarded = ''
        other_profile = ''
        profiles_info_pattern = re.compile("\s*([\s\S]*?)\s{\s}\n", re.MULTILINE)
        profiles_list = profiles_info_pattern.findall(vs_profiles)
        for profile in profiles_list:
            profile_name = profile.strip()
            if profile_name in ltm_v12_fastl4_profile_map.keys():
                fastl4_profile_name = profile_name
                fastl4_info_str = ltm_v12_fastl4_profile_map[profile_name]
                fastl4_timeout_pattern = re.compile("##timeout#([\s\S]*?)##", re.MULTILINE)
                fastl4_timeout = ''.join(fastl4_timeout_pattern.findall(fastl4_info_str))
                fastl4_pva_pattern = re.compile("##pva#([\s\S]*?)##", re.MULTILINE)
                fastl4_pva = ''.join(fastl4_pva_pattern.findall(fastl4_info_str))
            elif profile_name in ltm_v12_tcp_profile_map.keys():
                tcp_profile_name = profile_name
                tcp_profile_timeout = ltm_v12_tcp_profile_map[profile_name]
            elif profile_name in ltm_v12_http_profile_map.keys():
                http_profile_name = profile_name
                http_profile_xforwarded = ltm_v12_http_profile_map[profile_name]
            else:
                other_profile = other_profile + '\n' + profile_name

        vs[16] = fastl4_profile_name
        vs[17] = fastl4_timeout
        vs[18] = fastl4_pva
        vs[19] = tcp_profile_name
        vs[20] = tcp_profile_timeout
        vs[21] = http_profile_name
        vs[22] = http_profile_xforwarded
        vs[23] = other_profile.strip('\n')

        vs_rules_pattern = re.compile("\s*rules\s(none|{[\s\S]*?})\n", re.MULTILINE)
        vs_rules = ''.join(vs_rules_pattern.findall(vs_info))
        vs[24] = vs_rules

        vs_snat_pool_str_pattern = re.compile("\s*source-address-translation\s(none|{[\s\S]*?})\n", re.MULTILINE)
        vs_snat_pool_str = ''.join(vs_snat_pool_str_pattern.findall(vs_info))
        vs_snat_pool_name = 'none'
        if vs_snat_pool_str != 'none' and vs_snat_pool_str != '':
            vs_snat_pool_pattern = re.compile("\s*pool\s([\s\S]*?)\n", re.MULTILINE)
            vs_snat_pool_name = ''.join(vs_snat_pool_pattern.findall(vs_snat_pool_str))
        vs[25] = vs_snat_pool_name

        vs_source_port_pattern = re.compile("\s*source-port\s([\s\S]*?)\n", re.MULTILINE)
        vs_source_port = ''.join(vs_source_port_pattern.findall(vs_info))
        vs[26] = vs_source_port

        vs_vlans_pattern = re.compile("\s*vlans\s(none|{[\s\S]*?})\n", re.MULTILINE)
        vs_vlans = ''.join(vs_vlans_pattern.findall(vs_info))
        vs[27] = vs_vlans

        vs_profile_name = ''
        if fastl4_profile_name != '':
            vs_profile_name = fastl4_profile_name
        else:
            vs_profile_name = http_profile_name + ' ' + tcp_profile_name + ' ' + other_profile.strip('\n')

        vs_info_new = '##vs_name#'+name+'##vs_persist_name#'+vs_persist_name+'##vs_pool_name#'+vs_pool_name+'##vs_profile_name#'+vs_profile_name+'##vs_snat_pool_name#'+vs_snat_pool_name+'##'

        f5_vs_info_map[vs_ip_port_info_new] = vs_info_new

        ltm_v12_vs_list.append(vs)

    return ltm_v12_vs_list

nsae_real_map = {}
nsae_vs_device_map = {}
nsae_vs_sslhost_map = {}
def get_nsae_ssl_config(file_path,type,version,device_name):
    ssl_config_open = open(file_path, encoding='utf-8' ,errors='ignore')
    ssl_config_open_str = ssl_config_open.read()
    ssl_config_open.close()

    nsae_slb_real_map = {}
    nsae_slb_real_list = nsae_slb_real_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_real_list:
        nsae_slb_real_name = re.sub('\s*','',item[0])
        nsae_slb_real_ip = re.sub('\s*','',item[1])
        nsae_slb_real_port = re.sub('\s*','',item[2])
        nsae_slb_real_limit = re.sub('\s*','',item[3])
        nsae_slb_real_check = re.sub('\s*','',item[4])
        real_ip_port_str = nsae_slb_real_ip.strip() + ':' + nsae_slb_real_port.strip()
        nsae_slb_real_map[nsae_slb_real_name] = "##ipport#" + nsae_slb_real_ip + ':' + nsae_slb_real_port + '##limit#' + nsae_slb_real_limit + "##check#" + nsae_slb_real_check + '##'
        if not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$", real_ip_port_str):
            real_ip_port_str = re.sub(':0{1,}', ':', real_ip_port_str.lower())
        nsae_real_map[device_name + '_' + real_ip_port_str] = nsae_slb_real_name

    nsae_slb_real_disable_map = {}
    nsae_slb_real_disable_list = nsae_slb_real_disable_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_real_disable_list:
        nsae_slb_real_name = re.sub('\s*','',item)
        nsae_slb_real_disable_map[nsae_slb_real_name] = "disable"

    nsae_slb_group_member_map = {}
    nsae_slb_group_member_list = nsae_slb_group_member_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_group_member_list:
        nsae_slb_pool_name = re.sub('\s*','',item[0])
        nsae_slb_pool_member = re.sub('\s*','',item[1])
        nsae_slb_pool_info = nsae_slb_pool_member + '\n'
        if nsae_slb_pool_name in nsae_slb_group_member_map.keys():
            nsae_slb_pool_info = nsae_slb_pool_info + nsae_slb_group_member_map[nsae_slb_pool_name]

        nsae_slb_group_member_map[nsae_slb_pool_name] = nsae_slb_pool_info

    nsae_slb_policy_map = {}
    nsae_slb_policy_list = nsae_slb_policy_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_policy_list:
        nsae_slb_vs_name = re.sub('\s*','',item[0])
        nsae_slb_vs_pool = re.sub('\s*','',item[1])
        nsae_slb_policy_map[nsae_slb_vs_name] = nsae_slb_vs_pool

    nsae_ssl_host_map = {}
    nsae_ssl_host_list = nsae_ssl_host_pattern.findall(ssl_config_open_str)
    for item in nsae_ssl_host_list:
        nsae_ssl_host_name = re.sub('\s*','',item[0])
        nsae_slb_vs_name = re.sub('\s*','',item[1])
        nsae_ssl_host_map[nsae_slb_vs_name] = nsae_ssl_host_name

    nsae_slb_virtual_map = {}
    nsae_slb_virtual_list = nsae_slb_virtual_pattern.findall(ssl_config_open_str)
    for item in nsae_slb_virtual_list:
        nsae_slb_vs_name = re.sub('\s*','',item[0])
        nsae_slb_vs_ip = re.sub('\s*','',item[1])
        nsae_slb_vs_port = re.sub('\s*','',item[2])
        nsae_slb_vs_info = nsae_slb_vs_ip + ':' + nsae_slb_vs_port
        nsae_slb_virtual_map[nsae_slb_vs_name] = nsae_slb_vs_info
        nsae_slb_ssl_host = ''
        if nsae_slb_vs_name in nsae_ssl_host_map.keys():
            nsae_slb_ssl_host = nsae_ssl_host_map[nsae_slb_vs_name]

        vs_ip_port_str = nsae_slb_vs_ip.strip() + ':' + nsae_slb_vs_port.strip()
        if not re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[\s\S]*?$", vs_ip_port_str):
            vs_ip_port_str = re.sub(':0{1,}', ':', vs_ip_port_str.lower())

        nsae_vs_device_map[vs_ip_port_str] = device_name
        nsae_vs_sslhost_map[vs_ip_port_str] = device_name + ',' + nsae_slb_ssl_host + ',' + nsae_slb_vs_name

    nsae_ssl_vs_list = []
    for vs_name in nsae_slb_virtual_map.keys():
        nsae_ssl_vs_info = ['']*6
        nsae_ssl_vs_info[0] = nsae_slb_virtual_map[vs_name]
        nsae_ssl_vs_info[2] = vs_name
        nsae_ssl_vs_info[1] = ''
        if vs_name in nsae_ssl_host_map.keys():
            nsae_ssl_vs_info[1] = nsae_ssl_host_map[vs_name]
        nsae_ssl_vs_pool_name = ''
        nsae_ssl_vs_member_simple = ''
        nsae_ssl_vs_member_detail = ''
        if vs_name in nsae_slb_policy_map.keys():
            nsae_ssl_vs_pool_name = nsae_slb_policy_map[vs_name].strip()
            if nsae_ssl_vs_pool_name in nsae_slb_group_member_map.keys():
                real_members = nsae_slb_group_member_map[nsae_ssl_vs_pool_name].strip()
                real_members_list = real_members.split('\n')
                vs_member_detail_info = ''
                vs_member_simple_info = ''
                for real_member in real_members_list:
                    if real_member != '' and real_member is not None:
                        real_member_info = nsae_slb_real_map[real_member]
                        real_member_ipport_pattern = re.compile("##ipport#([\s\S]*?)##", re.MULTILINE)
                        real_member_ipport = ''.join(real_member_ipport_pattern.findall(real_member_info))
                        real_member_limit_pattern = re.compile("##limit#([\s\S]*?)##", re.MULTILINE)
                        real_member_limit = ''.join(real_member_limit_pattern.findall(real_member_info))
                        real_member_check_pattern = re.compile("##check#([\s\S]*?)##", re.MULTILINE)
                        real_member_check = ''.join(real_member_check_pattern.findall(real_member_info))
                        if real_member in nsae_slb_real_disable_map.keys():
                            vs_member_detail_info += real_member_ipport + ' disable' + ' l:' + real_member_limit + ' c:' + real_member_check + '\n'
                        else:
                            vs_member_detail_info += real_member_ipport + ' enable' + ' l:' + real_member_limit + ' c:' + real_member_check + '\n'
                            vs_member_simple_info += real_member_ipport + '\n'

                nsae_ssl_vs_member_simple = vs_member_simple_info.strip('\n')
                nsae_ssl_vs_member_detail = vs_member_detail_info.strip('\n')
            elif nsae_ssl_vs_pool_name in nsae_slb_real_map.keys():
                real_member_info = nsae_slb_real_map[nsae_ssl_vs_pool_name]
                real_member_ipport_pattern = re.compile("##ipport#([\s\S]*?)##", re.MULTILINE)
                real_member_ipport = ''.join(real_member_ipport_pattern.findall(real_member_info))
                real_member_limit_pattern = re.compile("##limit#([\s\S]*?)##", re.MULTILINE)
                real_member_limit = ''.join(real_member_limit_pattern.findall(real_member_info))
                real_member_check_pattern = re.compile("##check#([\s\S]*?)##", re.MULTILINE)
                real_member_check = ''.join(real_member_check_pattern.findall(real_member_info))
                vs_member_detail_info = real_member_ipport + ' enable' + ' l:' + real_member_limit + ' c:' + real_member_check
                vs_member_simple_info = ''
                if nsae_ssl_vs_pool_name in nsae_slb_real_disable_map.keys():
                    vs_member_detail_info = real_member_ipport + ' disable' + ' l:' + real_member_limit + ' c:' + real_member_check
                else:
                    vs_member_simple_info = vs_member_simple_info + real_member_ipport

                nsae_ssl_vs_member_simple = vs_member_simple_info
                nsae_ssl_vs_member_detail = vs_member_detail_info

        nsae_ssl_vs_info[3] = nsae_ssl_vs_pool_name
        nsae_ssl_vs_info[4] = nsae_ssl_vs_member_simple
        nsae_ssl_vs_info[5] = nsae_ssl_vs_member_detail
        nsae_ssl_vs_list.append(nsae_ssl_vs_info)


    return nsae_ssl_vs_list

def get_citrix_config(file_path, type, version):
    citrix_config_open = open(file_path, encoding='utf-8' ,errors='ignore')
    citrix_config_open_str = citrix_config_open.read()
    citrix_config_open.close()

    citrix_server_map = {}
    citrix_servers = citrix_server_pattern.findall(citrix_config_open_str)
    for item in citrix_servers:
        server_name = item[0].strip()
        server_ip = item[1].strip()
        citrix_server_map[server_name] = server_ip

    citrix_pool_mem_map = {}
    citrix_pool_mems = citrix_pool_mem_pattern.findall(citrix_config_open_str)
    for item in citrix_pool_mems:
        pool_name = item[0].strip()
        pool_name_mem = pool_name + '_mem'
        pool_name_mon = pool_name + '_mon'
        mem_str = item[1]
        if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}\s\d*$", mem_str):
            ip_port_str = mem_str.replace(' ', ':')
            if pool_name_mem in citrix_pool_mem_map.keys():
                tmp = citrix_pool_mem_map[pool_name_mem]
                citrix_pool_mem_map[pool_name_mem] = tmp + ip_port_str
            else:
                citrix_pool_mem_map[pool_name_mem] = ip_port_str
            citrix_pool_mem_map[pool_name_mon] = ''
        else:
            if "monitorName" in mem_str:
                mon_pattern = re.compile("monitorName\s([\s\S]*?)\n", re.MULTILINE)
                mon = ''.join(mon_pattern.findall(mem_str))
                citrix_pool_mem_map[pool_name_mon] = mon
            else:
                server_ports = mem_str.strip().split(" ")
                server = server_ports[0].strip()
                port = server_ports[1].strip()
                ip_port_str = ''
                if server in citrix_server_map.keys():
                    ip_port_str = citrix_server_map[server] + ':' + port + '\n'
                else:
                    ip_port_str = server + ':' + port + '\n'

                if pool_name_mem in citrix_pool_mem_map.keys():
                    tmp = citrix_pool_mem_map[pool_name_mem]
                    citrix_pool_mem_map[pool_name_mem] = tmp + ip_port_str
                else:
                    citrix_pool_mem_map[pool_name_mem] = ip_port_str

                citrix_pool_mem_map[pool_name_mon] = ''

    citrix_pool_map = {}
    citrix_pools = citrix_pool_pattern.findall(citrix_config_open_str)
    for item in citrix_pools:
        pool_name = item[0].strip()
        pool_protocol = item[1].strip()
        pool_info_str = item[2].strip()
        is_xff = " "
        if "X-Forwarded-For" in pool_info_str:
            is_xff = 'enabled'
        pool_mems = citrix_pool_mem_map[pool_name + '_mem']
        pool_mon = citrix_pool_mem_map[pool_name + '_mon']
        pool_info = '##pool_name#' + pool_name +'##pool_protocol#' + pool_protocol + '##is_xff#'+ is_xff + '##pool_mems#' + pool_mems + '##pool_mon#' + pool_mon + "##"
        citrix_pool_map[pool_name] = pool_info

    citrix_policy_map = {}
    citrix_policys = citrix_policy_pattern.findall(citrix_config_open_str)
    for item1 in citrix_policys:
        policy_name = item1[0].strip()
        policy_info_str1 = item1[1].strip()
        domain_name_pattern = re.compile('HTTP.REQ.HOSTNAME.EQ\(([\s\S]*?)\)', re.MULTILINE)
        domain_name = '\n'.join(domain_name_pattern.findall(policy_info_str1)).replace('\\"','')
        citrix_policy_map[policy_name] = '##domain_name#' + domain_name + '##'

    citrix_vs_policy_map = {}
    citrix_vs_policys = citrix_vs_policy_pattern.findall(citrix_config_open_str)
    for item in citrix_vs_policys:
        vs_name = item[0].strip()
        policy_info_str = item[1].strip()
        policy_name_pattern = re.compile("-policyName\s([\s\S]*?)\s", re.MULTILINE)
        policy_name = ''.join(policy_name_pattern.findall(policy_info_str))
        policy_info = ''
        pool_info = ''
        if policy_name == '':
            pool_info = citrix_pool_map[policy_info_str]
            if vs_name in citrix_vs_policy_map.keys():
                citrix_vs_policy_map[vs_name] += pool_info
            else:
                citrix_vs_policy_map[vs_name] = pool_info
        else:
            policy_info = citrix_policy_map[policy_name]
            if vs_name in citrix_vs_policy_map.keys():
                citrix_vs_policy_map[vs_name] += policy_info
            else:
                citrix_vs_policy_map[vs_name] = policy_info

    citrix_vs_list = []
    citrix_vss = citrix_vs_pattern.findall(citrix_config_open_str)
    for item in citrix_vss:
        vs = [''] * 14
        vs_name = item[0].strip()
        vs[0] = vs_name
        vs_protocol= item[1].strip()
        vs[1] = vs_protocol
        vs_ip = item[2].strip()
        vs_port = item[3].strip()
        vs[2] = vs_ip + ':' + vs_port
        persist_type = item[4].strip()
        vs[5] = persist_type
        status_info_str = item[5].strip()

        vs_status = 'enabled'
        if re.match(r"-state\sDISABLED", status_info_str):
            vs_status = 'disabled'
        vs[3] = vs_status

        vs_tcp_timeout_pattern = re.compile("-cltTimeout\s(\d*)", re.MULTILINE)
        vs_tcp_timeout = ''.join(vs_tcp_timeout_pattern.findall(status_info_str))
        vs[4] = vs_tcp_timeout

        vs_persist_timeout_pattern = re.compile("-timeout\s(\d*)", re.MULTILINE)
        vs_persist_timeout = ''.join(vs_persist_timeout_pattern.findall(status_info_str))
        vs[6] = vs_persist_timeout

        vs_snatpool = item[6].strip()
        vs[12] = vs_snatpool

        vs_policy_info_str = citrix_vs_policy_map[vs_name]

        pool_name_pattern = re.compile("##pool_name#([\s\S]*?)##", re.MULTILINE)
        pool_name = ''.join(pool_name_pattern.findall(vs_policy_info_str))
        vs[7] = pool_name

        pool_protocol_pattern = re.compile("##pool_protocol#([\s\S]*?)##", re.MULTILINE)
        pool_protocol = ''.join(pool_protocol_pattern.findall(vs_policy_info_str))
        vs[8] = pool_protocol

        is_xff_pattern = re.compile("##is_xff#([\s\S]*?)##", re.MULTILINE)
        is_xff = ''.join(is_xff_pattern.findall(vs_policy_info_str))
        vs[11] = is_xff

        pool_mems_pattern = re.compile("##pool_mems#([\s\S]*?)##", re.MULTILINE)
        pool_mems = ''.join(pool_mems_pattern.findall(vs_policy_info_str))
        vs[9] = pool_mems.strip('\n')

        pool_mon_pattern = re.compile("##pool_mon#([\s\S]*?)##", re.MULTILINE)
        pool_mon = ''.join(pool_mon_pattern.findall(vs_policy_info_str))
        vs[10] = pool_mon

        domain_name_pattern = re.compile("##domain_name#([\s\S]*?)##", re.MULTILINE)
        domain_name = ''.join(domain_name_pattern.findall(vs_policy_info_str))
        vs[13] = domain_name

        citrix_vs_list.append(vs)

    return citrix_vs_list
def main():
    get_device_list()
    now_time = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    ltm_writer = pd.ExcelWriter(config_path_os + 'ltm_'+now_time+'.xlsx')
    nsae_writer = pd.ExcelWriter(config_path_os + 'nsae_'+now_time+'.xlsx')
    citrix_writer = pd.ExcelWriter(config_path_os + 'citrix_'+now_time+'.xlsx')
    # gtm_writer = pd.ExcelWriter(config_path_os + 'gtm_'+now_time+'.xlsx')
    device_base_writer = pd.ExcelWriter(config_path_os + 'device_base_' + now_time + '.xlsx')

    device_base_list = []
    for device_name in device_analyzer_list:
        file_path = device_path_map[device_name]
        type = device_list_map[device_name]['type']
        version = device_list_map[device_name]['version']
        # print(file_path)
        if type == 'ltm':
            ltm_list = get_ltm_config(file_path,type,version,device_name)
            ltm_df = pd.DataFrame(ltm_list, columns=['vs_name', 'vs_conn', 'vs_ip_port', 'vs_status', 'vs_protocol', 'vs_persist_name', 'vs_persist_mothod', 'vs_persist_timeout', 'persist_cookie_encrypt', 'persist_cookie_name', 'persist_cookie_method', 'vs_pool_name', 'vs_balanc_mode', 'vs_pool_monitor', 'members_info_simple', 'members_info_detail', 'fastl4_profile_name', 'fastl4_timeout', 'fastl4_pva', 'tcp_profile_name', 'tcp_profile_timeout', 'http_profile_name', 'http_profile_xforwarded', 'other_profile', 'vs_rules', 'vs_snat_pool_name', 'vs_source_port', 'vs_vlans'])
            ltm_df.to_excel(ltm_writer, sheet_name=device_name, index=False)
        elif type == 'nsae':
            nsae_ssl_vs_list = get_nsae_ssl_config(file_path, type, version,device_name)
            nsae_df = pd.DataFrame(nsae_ssl_vs_list, columns=['ssl_vs_ipport', 'ssl_host', 'ssl_vs_name', 'ssl_pool_name', 'members_info_simple', 'members_info_detail'])
            nsae_df.to_excel(nsae_writer, sheet_name=device_name, index=False)
        elif type == 'citrix':
            citrix_vs_list = get_citrix_config(file_path, type, version)
            citrix_df = pd.DataFrame(citrix_vs_list, columns=['vs_name', 'vs_protocol', 'vs_ip_port', 'vs_status', 'vs_tcp_timeout', 'persist_type', 'persist_time', 'pool_name', 'pool_protocol', 'pool_mems', 'pool_mon', 'xforwarded', 'vs_snatpool', 'domain_name'])
            citrix_df.to_excel(citrix_writer, sheet_name=device_name, index=False)
        # elif type == 'gtm':
            # gtm_list = []
            # gtm = [''] * 4
            # gtm[0] = device_name
            # gtm[1] = type
            # gtm[2] = device_list_map[device_name]['version']
            # gtm[3] = device_path_map[device_name]
            # gtm_list.append(gtm)
            # gtm_df = pd.DataFrame(gtm_list, columns=['device_name', 'type', 'version', 'path'])
            # gtm_df.to_excel(gtm_writer, sheet_name=device_name, index=False)

        ltm_base_config = get_ltm_base_config(file_path,type,version,device_name)
        device_base_list.append(ltm_base_config)

    device_base_df = pd.DataFrame(device_base_list,
                          columns=['device_name', 'routes', 'self_ips', 'acls', 'ntp', 'snmp',
                                   'syslog', 'snatpool'])
    device_base_df.to_excel(device_base_writer, sheet_name='device_base', index=False)

    ltm_writer.close()
    nsae_writer.close()
    citrix_writer.close()
    # # gtm_writer.close()
    device_base_writer.close()

    get_waf_pass_list()

    ccvcc_waf_pass_writer = pd.ExcelWriter(config_path_os + '变更_网站旁路waf_'+now_time+'.xlsx')
    ccvec_waf_pass_writer = pd.ExcelWriter(config_path_os + '变更_对公旁路waf_' + now_time + '.xlsx')
    ccvep_waf_pass_writer = pd.ExcelWriter(config_path_os + '变更_对私旁路waf_'+now_time+'.xlsx')
    ccvmb_waf_pass_writer = pd.ExcelWriter(config_path_os + '变更_手机旁路waf_'+now_time+'.xlsx')

    for domain in waf_paas_map.keys():
        newlist = waf_paas_map[domain].strip("\n").split("\n")
        newlist.sort()
        scripts_list = []
        system_str = ''
        for item in newlist:
            scripts = item.split("##")
            system_str = scripts[1]
            device_name = scripts[2]
            script_strs = scripts[4]
            script1 = [device_name,script_strs]
            scripts_list.append(script1)

        if '网站' in system_str:
            ccvcc_waf_pass_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvcc_waf_pass_df.to_excel(ccvcc_waf_pass_writer, sheet_name=domain, index=False)
        elif '对公' in system_str:
            ccvec_waf_pass_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvec_waf_pass_df.to_excel(ccvec_waf_pass_writer, sheet_name=domain, index=False)
        elif '对私' in system_str:
            ccvep_waf_pass_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvep_waf_pass_df.to_excel(ccvep_waf_pass_writer, sheet_name=domain, index=False)
        elif '手机' in system_str:
            ccvmb_waf_pass_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvmb_waf_pass_df.to_excel(ccvmb_waf_pass_writer, sheet_name=domain, index=False)



    ccvcc_waf_pass_writer.close()
    ccvec_waf_pass_writer.close()
    ccvep_waf_pass_writer.close()
    ccvmb_waf_pass_writer.close()

    ccvcc_waf_over_writer = pd.ExcelWriter(config_path_os + '回退_网站旁路waf_' + now_time + '.xlsx')
    ccvec_waf_over_writer = pd.ExcelWriter(config_path_os + '回退_对公旁路waf_' + now_time + '.xlsx')
    ccvep_waf_over_writer = pd.ExcelWriter(config_path_os + '回退_对私旁路waf_' + now_time + '.xlsx')
    ccvmb_waf_over_writer = pd.ExcelWriter(config_path_os + '回退_手机旁路waf_' + now_time + '.xlsx')

    for domain in waf_over_map.keys():
        newlist = waf_over_map[domain].strip("\n").split("\n")
        newlist.sort()
        scripts_list = []
        system_str = ''
        for item in newlist:
            scripts = item.split("##")
            system_str = scripts[1]
            device_name = scripts[2]
            script_strs = scripts[4]
            script1 = [device_name,script_strs]
            scripts_list.append(script1)

        if '网站' in system_str:
            ccvcc_waf_over_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvcc_waf_over_df.to_excel(ccvcc_waf_over_writer, sheet_name=domain, index=False)
        elif '对公' in system_str:
            ccvec_waf_over_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvec_waf_over_df.to_excel(ccvec_waf_over_writer, sheet_name=domain, index=False)
        elif '对私' in system_str:
            ccvep_waf_over_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvep_waf_over_df.to_excel(ccvep_waf_over_writer, sheet_name=domain, index=False)
        elif '手机' in system_str:
            ccvmb_waf_over_df = pd.DataFrame(scripts_list, columns=['device_name', 'commad'])
            ccvmb_waf_over_df.to_excel(ccvmb_waf_over_writer, sheet_name=domain, index=False)

    ccvcc_waf_over_writer.close()
    ccvec_waf_over_writer.close()
    ccvep_waf_over_writer.close()
    ccvmb_waf_over_writer.close()

    ccvcc_sslupdate_writer = pd.ExcelWriter(config_path_os + 'sslupdate_网站_' + now_time + '.xlsx')
    ccvec_sslupdate_writer = pd.ExcelWriter(config_path_os + 'sslupdate_对公_' + now_time + '.xlsx')
    ccvep_sslupdate_writer = pd.ExcelWriter(config_path_os + 'sslupdate_对私_' + now_time + '.xlsx')
    ccvmb_sslupdate_writer = pd.ExcelWriter(config_path_os + 'sslupdate_手机_' + now_time + '.xlsx')

    bak_time = datetime.datetime.now().strftime('%Y%m%d')

    for domain in ssl_update_map.keys():
        newlist = ssl_update_map[domain].strip("\n").split("\n")
        newlist.sort()
        scripts_list = []
        bak_scripts_list = []
        restore_scripts_list = []
        system_str = ''
        for item in newlist:
            scripts = item.split("##")
            system_str = scripts[1]
            script_strs = scripts[2].split(",")
            script1 = [domain, script_strs[0], script_strs[1], script_strs[2]]
            scripts_list.append(script1)
            bak_script = 'ssl backup certificate "' + script_strs[1] + '" "bak'+ bak_time + '" ' + '"123123" '
            bak_scripts = [script_strs[0], bak_script]
            bak_scripts_list.append(bak_scripts)

            restore_script = 'ssl restore certificate "' + script_strs[1] + '" "bak'+ bak_time + '" ' + '"123123" '
            restore_scripts = [script_strs[0], restore_script]
            restore_scripts_list.append(restore_scripts)

        if '网站' in system_str:
            ccvcc_sslupdate_df = pd.DataFrame(scripts_list, columns=['domain', 'device_name','ssl_host','vs_name'])
            ccvcc_sslupdate_df.to_excel(ccvcc_sslupdate_writer, sheet_name=domain, index=False)
            ccvcc_bak_df = pd.DataFrame(bak_scripts_list, columns=['device_name','bak_scripts'])
            ccvcc_bak_df.to_excel(ccvcc_sslupdate_writer, sheet_name=domain+'_bak', index=False)
            ccvcc_restore_df = pd.DataFrame(restore_scripts_list, columns=['device_name','restore_scripts'])
            ccvcc_restore_df.to_excel(ccvcc_sslupdate_writer, sheet_name=domain+'_restore', index=False)
        elif '对公' in system_str:
            ccvec_sslupdate_df = pd.DataFrame(scripts_list, columns=['domain', 'device_name','ssl_host','vs_name'])
            ccvec_sslupdate_df.to_excel(ccvec_sslupdate_writer, sheet_name=domain, index=False)
            ccvec_bak_df = pd.DataFrame(bak_scripts_list, columns=['device_name','bak_scripts'])
            ccvec_bak_df.to_excel(ccvec_sslupdate_writer, sheet_name=domain+'_bak', index=False)
            ccvec_restore_df = pd.DataFrame(restore_scripts_list, columns=['device_name','restore_scripts'])
            ccvec_restore_df.to_excel(ccvec_sslupdate_writer, sheet_name=domain+'_restore', index=False)
        elif '对私' in system_str:
            ccvep_sslupdate_df = pd.DataFrame(scripts_list, columns=['domain', 'device_name','ssl_host','vs_name'])
            ccvep_sslupdate_df.to_excel(ccvep_sslupdate_writer, sheet_name=domain, index=False)
            ccvep_bak_df = pd.DataFrame(bak_scripts_list, columns=['device_name','bak_scripts'])
            ccvep_bak_df.to_excel(ccvep_sslupdate_writer, sheet_name=domain+'_bak', index=False)
            ccvep_restore_df = pd.DataFrame(restore_scripts_list, columns=['device_name','restore_scripts'])
            ccvep_restore_df.to_excel(ccvep_sslupdate_writer, sheet_name=domain+'_restore', index=False)
        elif '手机' in system_str:
            ccvmb_sslupdate_df = pd.DataFrame(scripts_list, columns=['domain', 'device_name','ssl_host','vs_name'])
            ccvmb_sslupdate_df.to_excel(ccvmb_sslupdate_writer, sheet_name=domain, index=False)
            ccvmb_bak_df = pd.DataFrame(bak_scripts_list, columns=['device_name','bak_scripts'])
            ccvmb_bak_df.to_excel(ccvmb_sslupdate_writer, sheet_name=domain+'_bak', index=False)
            ccvmb_restore_df = pd.DataFrame(restore_scripts_list, columns=['device_name','restore_scripts'])
            ccvmb_restore_df.to_excel(ccvmb_sslupdate_writer, sheet_name=domain+'_restore', index=False)

    ccvcc_sslupdate_writer.close()
    ccvec_sslupdate_writer.close()
    ccvep_sslupdate_writer.close()
    ccvmb_sslupdate_writer.close()


if __name__ == '__main__':
    main()






